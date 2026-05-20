"""Genera la plantilla Excel onboarding_template.xlsx
Lo correrías UNA vez para regenerar la plantilla; el archivo .xlsx queda
versionado y se le entrega al cliente para que lo llene."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUTPUT = "onboarding_template.xlsx"

wb = Workbook()

# ============================================================================
# Sheet 1: Instructions
# ============================================================================
ws_help = wb.active
ws_help.title = "Instrucciones"

instructions = [
    ("Plantilla de onboarding - Falabella ValueData", True, 16),
    ("", False, 11),
    ("Esta planilla sirve para sumar usuarios al canal WhatsApp del POC.", False, 11),
    ("Cada fila representa una persona que va a recibir o usar el bot.", False, 11),
    ("", False, 11),
    ("Pasos:", True, 12),
    ("1. Ir a la pestaña 'Personas' y llenar una fila por persona.", False, 11),
    ("2. Completar las columnas obligatorias (marcadas con asterisco).", False, 11),
    ("3. Guardar el archivo y enviarlo a Value Data.", False, 11),
    ("4. Value Data corre el script bulk_onboard.py contra esta planilla.", False, 11),
    ("5. El script crea los usuarios en el sistema, genera un wa.me link por cada uno", False, 11),
    ("   y lo escribe de vuelta en la columna 'Activation Link'.", False, 11),
    ("6. Se entrega la planilla actualizada al cliente.", False, 11),
    ("7. Cada persona hace click en su link, manda el mensaje pre-rellenado,", False, 11),
    ("   y queda activa para conversar con el bot por WhatsApp.", False, 11),
    ("", False, 11),
    ("Tipos de Persona (columna 'Tipo'):", True, 12),
    ("• user      - Acceso a la app web (admin, ops, jefe de transporte).", False, 11),
    ("• driver    - Conductor de un vehículo. Solo recibe/manda por WhatsApp.", False, 11),
    ("• contacto  - Notificaciones extra por empresa. No accede a la app.", False, 11),
    ("", False, 11),
    ("Columnas obligatorias por Tipo:", True, 12),
    ("• user:      Tipo, Nombre, Email, Telefono, Rol Usuario, Password", False, 11),
    ("• driver:    Tipo, Driver ID, Nombre, Telefono, Empresa ID, Vehiculo ID, Vehiculo Nombre", False, 11),
    ("• contacto:  Tipo, Nombre, Telefono, Empresa ID, Rol Contacto, Severities", False, 11),
    ("", False, 11),
    ("Otras notas:", True, 12),
    ("• El 'Telefono' es siempre en formato E.164: +56912345678 (sin espacios).", False, 11),
    ("• 'Empresa ID' es el numero interno; pedirlo al admin si no lo saben.", False, 11),
    ("• La columna 'Activation Link' la llena el script, dejarla vacia al armar.", False, 11),
    ("• 'Notas' es libre, solo para que el cliente identifique al usuario.", False, 11),
]

for i, (text, bold, size) in enumerate(instructions, start=1):
    cell = ws_help.cell(row=i, column=1, value=text)
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

ws_help.column_dimensions["A"].width = 110

# ============================================================================
# Sheet 2: Personas (data)
# ============================================================================
ws = wb.create_sheet("Personas")

headers = [
    ("Tipo*", "user / driver / contacto"),
    ("Nombre*", "Nombre completo de la persona"),
    ("Email", "Solo para 'user'. Identifica el login en la app web."),
    ("Telefono*", "E.164 (ej. +56912345678). Sin espacios ni guiones."),
    ("Empresa ID", "Solo para 'driver' y 'contacto'. Numero interno de la empresa de transporte."),
    ("Rol Usuario", "Solo para 'user'. Valores: falabella_admin, falabella_ops, transport_manager."),
    ("Password", "Solo para 'user'. Minimo 4 caracteres. Idealmente cambiar despues del primer login."),
    ("Driver ID", "Solo para 'driver'. Codigo unico, ej. DRV-020. Si esta vacio se autogenera."),
    ("Vehiculo ID", "Solo para 'driver'. ID del vehiculo asignado."),
    ("Vehiculo Nombre", "Solo para 'driver'. Ej. FAL-1015."),
    ("Rol Contacto", "Solo para 'contacto'. Valores: coordinador, dispatcher, jefe, otro."),
    ("Severities", "Solo para 'contacto'. Lista separada por coma: critical,high,medium,low. Si vacia, recibe todas."),
    ("Notas", "Libre. Por ejemplo, area o cargo de la persona."),
    ("Activation Link", "<- NO LLENAR, lo llena el script automaticamente despues."),
]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2A3F5F", end_color="2A3F5F", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style="thin", color="BBBBBB"),
    right=Side(style="thin", color="BBBBBB"),
    top=Side(style="thin", color="BBBBBB"),
    bottom=Side(style="thin", color="BBBBBB"),
)

for col_idx, (label, comment_text) in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    cell.comment = Comment(comment_text, "ValueData")
    ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(label) + 2)

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 28
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 22
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 14
ws.column_dimensions["I"].width = 12
ws.column_dimensions["J"].width = 16
ws.column_dimensions["K"].width = 16
ws.column_dimensions["L"].width = 32
ws.column_dimensions["M"].width = 28
ws.column_dimensions["N"].width = 60

# Data validations
dv_tipo = DataValidation(
    type="list",
    formula1='"user,driver,contacto"',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle="Tipo invalido",
    error="Elegi entre user, driver o contacto.",
)
dv_tipo.add("A2:A1000")
ws.add_data_validation(dv_tipo)

dv_rol_user = DataValidation(
    type="list",
    formula1='"falabella_admin,falabella_ops,transport_manager"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Rol invalido",
    error="Para users, los roles validos son falabella_admin, falabella_ops o transport_manager.",
)
dv_rol_user.add("F2:F1000")
ws.add_data_validation(dv_rol_user)

dv_rol_contacto = DataValidation(
    type="list",
    formula1='"coordinador,dispatcher,jefe,otro"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Rol contacto invalido",
    error="Para contactos, los roles validos son coordinador, dispatcher, jefe u otro.",
)
dv_rol_contacto.add("K2:K1000")
ws.add_data_validation(dv_rol_contacto)

# Ejemplos
example_rows = [
    [
        "user", "Maria Gonzalez", "maria@cliente.cl", "+56911112222",
        "", "transport_manager", "Cambio123", "", "", "", "", "",
        "Jefa de operacion de transporte", ""
    ],
    [
        "driver", "Juan Perez", "", "+56933334444",
        22, "", "", "DRV-020", 1, "FAL-1000", "", "",
        "Driver del turno tarde", ""
    ],
    [
        "contacto", "Pedro Lopez", "", "+56955556666",
        22, "", "", "", "", "", "jefe", "critical,high",
        "Manager operativo, solo alertas urgentes", ""
    ],
]

for r_idx, row in enumerate(example_rows, start=2):
    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.font = Font(italic=True, color="888888")  # ejemplo en gris cursiva

# Highlight example header row
ws.cell(row=2, column=1).comment = Comment(
    "Esta es una fila de EJEMPLO. Borrala antes de cargar tus propias filas.",
    "ValueData",
)

ws.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"OK: {OUTPUT} creado.")
