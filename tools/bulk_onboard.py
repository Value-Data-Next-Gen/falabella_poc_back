"""Procesa onboarding_template.xlsx llenado y crea users/drivers/contactos
contra el backend de Falabella ValueData. Devuelve la planilla con la
columna 'Activation Link' poblada para cada fila exitosa.

Uso:
    pip install openpyxl requests
    python bulk_onboard.py onboarding_template.xlsx \\
        --api-base https://poc-fal-back-...azurewebsites.net \\
        --admin-email admin@falabella.cl --admin-password admin123

Opciones útiles:
    --dry-run         : valida sin crear nada
    --output FILE     : ruta del Excel actualizado (default: <input>_done.xlsx)
"""
from __future__ import annotations

import argparse
import sys
import time
from pathlib import Path

try:
    import requests
except ImportError:
    print("Falta 'requests'. Instala: pip install requests openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("Falta 'openpyxl'. Instala: pip install requests openpyxl", file=sys.stderr)
    sys.exit(1)


GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")


def login(api_base: str, email: str, password: str) -> str:
    r = requests.post(
        f"{api_base.rstrip('/')}/api/auth/login",
        json={"email": email, "password": password},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()["access_token"]


def post_user(api_base: str, token: str, row: dict) -> dict:
    body = {
        "email": (row.get("Email") or "").strip(),
        "password": (row.get("Password") or "").strip(),
        "display_name": (row.get("Nombre*") or row.get("Nombre") or "").strip(),
        "role": (row.get("Rol Usuario") or "").strip(),
        "phone_e164": (row.get("Telefono*") or row.get("Telefono") or "").strip(),
        "notify_whatsapp": True,
        "activo": True,
    }
    r = requests.post(
        f"{api_base.rstrip('/')}/api/admin/users",
        json=body,
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()


def post_driver(api_base: str, token: str, row: dict) -> dict:
    body = {
        "driver_id": (row.get("Driver ID") or "").strip(),
        "name": (row.get("Nombre*") or row.get("Nombre") or "").strip(),
        "phone": (row.get("Telefono*") or row.get("Telefono") or "").strip(),
        "empresa_id": int(row.get("Empresa ID")) if row.get("Empresa ID") else None,
        "vehicle_id": int(row.get("Vehiculo ID")) if row.get("Vehiculo ID") else None,
        "vehicle_name": (row.get("Vehiculo Nombre") or "").strip(),
        "license": "A-3 Profesional",
        "rating": 4.5,
        "active": True,
    }
    if not body["driver_id"]:
        body["driver_id"] = f"DRV-{int(time.time() * 1000) % 1_000_000:06d}"
    r = requests.post(
        f"{api_base.rstrip('/')}/api/admin/drivers",
        json=body,
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()


def post_contacto(api_base: str, token: str, row: dict) -> dict:
    severities_raw = (row.get("Severities") or "").strip()
    severities = [s.strip() for s in severities_raw.split(",") if s.strip()] if severities_raw else None
    empresa_id = int(row["Empresa ID"])
    body = {
        "nombre": (row.get("Nombre*") or row.get("Nombre") or "").strip(),
        "rol": (row.get("Rol Contacto") or "otro").strip(),
        "phone_e164": (row.get("Telefono*") or row.get("Telefono") or "").strip(),
        "region_filter": "all",
    }
    if severities:
        body["severities_in"] = severities
    r = requests.post(
        f"{api_base.rstrip('/')}/api/empresa-contactos/empresas/{empresa_id}/contactos",
        json=body,
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()


def get_activation_link(api_base: str, token: str, kind: str, id_field) -> str | None:
    headers = {"Authorization": f"Bearer {token}"}
    base = api_base.rstrip("/")
    if kind == "user":
        url = f"{base}/api/admin/users/{id_field}/activation-link"
    elif kind == "driver":
        url = f"{base}/api/admin/drivers/{id_field}/activation-link"
    elif kind == "contacto":
        empresa_id, contact_id = id_field
        url = f"{base}/api/empresa-contactos/empresas/{empresa_id}/contactos/{contact_id}/activation-link"
    else:
        return None
    r = requests.get(url, headers=headers, timeout=30)
    if r.status_code != 200:
        return None
    return r.json().get("link")


def validate_row(row: dict) -> str | None:
    tipo = (row.get("Tipo*") or "").strip().lower()
    nombre = (row.get("Nombre*") or "").strip()
    phone = (row.get("Telefono*") or "").strip()
    if not tipo or tipo not in ("user", "driver", "contacto"):
        return "Tipo invalido (debe ser user/driver/contacto)"
    if not nombre:
        return "Nombre vacio"
    if not phone or not phone.startswith("+"):
        return "Telefono vacio o no esta en formato E.164 (+...)"
    if tipo == "user":
        if not (row.get("Email") or "").strip():
            return "User requiere Email"
        if not (row.get("Password") or "").strip():
            return "User requiere Password"
        if (row.get("Rol Usuario") or "").strip() not in ("falabella_admin", "falabella_ops", "transport_manager"):
            return "Rol Usuario invalido"
    elif tipo == "driver":
        for f in ("Empresa ID", "Vehiculo ID", "Vehiculo Nombre"):
            if not row.get(f):
                return f"Driver requiere {f}"
    elif tipo == "contacto":
        if not row.get("Empresa ID"):
            return "Contacto requiere Empresa ID"
        rol = (row.get("Rol Contacto") or "").strip()
        if rol not in ("coordinador", "dispatcher", "jefe", "otro", ""):
            return "Rol Contacto invalido"
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", help="Path al onboarding_template.xlsx lleno")
    parser.add_argument("--api-base", default="https://poc-fal-back-hrgccdc4hub8fuea.brazilsouth-01.azurewebsites.net")
    parser.add_argument("--admin-email", default="admin@falabella.cl")
    parser.add_argument("--admin-password", default="admin123")
    parser.add_argument("--output", default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    xlsx_in = Path(args.xlsx)
    if not xlsx_in.exists():
        print(f"No existe {xlsx_in}", file=sys.stderr)
        sys.exit(1)
    xlsx_out = Path(args.output) if args.output else xlsx_in.with_name(xlsx_in.stem + "_done.xlsx")

    print(f"Login a {args.api_base}...")
    token = login(args.api_base, args.admin_email, args.admin_password)
    print(f"OK\n")

    wb = load_workbook(xlsx_in)
    if "Personas" not in wb.sheetnames:
        print("La planilla no tiene la pestana 'Personas'.", file=sys.stderr)
        sys.exit(1)
    ws = wb["Personas"]

    # Locate columns
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}
    if "Activation Link" not in col_idx:
        print("Falta columna 'Activation Link' en la planilla.", file=sys.stderr)
        sys.exit(1)
    link_col = col_idx["Activation Link"]

    summary = {"ok": 0, "skipped": 0, "error": 0}
    for r_idx in range(2, ws.max_row + 1):
        row_dict = {h: ws.cell(row=r_idx, column=col_idx[h]).value for h in col_idx}
        # Skip empty rows
        if not any((row_dict.get(h) or "") for h in ("Tipo*", "Nombre*", "Telefono*")):
            continue
        # Skip example rows (italic gray font)
        first_cell = ws.cell(row=r_idx, column=1)
        if first_cell.font and first_cell.font.italic and (first_cell.font.color and str(first_cell.font.color.rgb).endswith("888888")):
            print(f"Row {r_idx}: skip ejemplo")
            summary["skipped"] += 1
            continue

        err = validate_row(row_dict)
        if err:
            print(f"Row {r_idx}: validacion fallida - {err}")
            ws.cell(row=r_idx, column=link_col, value=f"ERROR: {err}").fill = RED
            summary["error"] += 1
            continue

        tipo = row_dict["Tipo*"].strip().lower()
        nombre = row_dict["Nombre*"].strip()

        if args.dry_run:
            print(f"Row {r_idx}: DRY-RUN {tipo} {nombre}")
            ws.cell(row=r_idx, column=link_col, value="(dry-run)").fill = YELLOW
            continue

        try:
            if tipo == "user":
                created = post_user(args.api_base, token, row_dict)
                link = get_activation_link(args.api_base, token, "user", created["user_id"])
            elif tipo == "driver":
                created = post_driver(args.api_base, token, row_dict)
                link = get_activation_link(args.api_base, token, "driver", created["driver_id"])
            elif tipo == "contacto":
                created = post_contacto(args.api_base, token, row_dict)
                link = get_activation_link(
                    args.api_base, token, "contacto",
                    (created["empresa_id"], created["contact_id"]),
                )
            else:
                continue
        except requests.HTTPError as e:
            detail = ""
            try:
                detail = e.response.json().get("detail", "")
            except Exception:
                detail = e.response.text[:200] if e.response is not None else ""
            print(f"Row {r_idx}: HTTP {e.response.status_code if e.response else '?'} - {detail}")
            ws.cell(row=r_idx, column=link_col, value=f"HTTP error: {detail[:150]}").fill = RED
            summary["error"] += 1
            continue
        except Exception as e:  # noqa: BLE001
            print(f"Row {r_idx}: exception - {e}")
            ws.cell(row=r_idx, column=link_col, value=f"exception: {str(e)[:150]}").fill = RED
            summary["error"] += 1
            continue

        cell = ws.cell(row=r_idx, column=link_col, value=link or "(sin link)")
        cell.fill = GREEN
        if link:
            cell.hyperlink = link
            cell.font = Font(color="0000EE", underline="single")
        print(f"Row {r_idx}: OK {tipo} {nombre} -> {link}")
        summary["ok"] += 1

    wb.save(xlsx_out)
    print(f"\n--- Resumen ---")
    print(f"OK:       {summary['ok']}")
    print(f"Skipped:  {summary['skipped']}")
    print(f"Errores:  {summary['error']}")
    print(f"Salida:   {xlsx_out}")


if __name__ == "__main__":
    main()
