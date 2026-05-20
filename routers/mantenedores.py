"""Mantenedor de maestros: CRUD admin sobre empresas, users, drivers, vehicles, clients.

Todos los endpoints requieren rol `falabella_admin`. Tras cada mutación de
drivers/vehicles/clients se llama `STATE.reload_maestros()` para que el cache
in-memory quede consistente con la DB.

Endpoints (prefijo /api/admin):
  Empresas:  GET, POST, PUT/{id}, DELETE/{id}     (/empresas)
  Users:     GET, POST, PUT/{id}, DELETE/{id}, POST/{id}/reset-password (/users)
  Drivers:   GET, POST, PUT/{id}, DELETE/{id}     (/drivers)
  Vehicles:  GET, POST, PUT/{id}, DELETE/{id}     (/vehicles)
  Clients:   GET (paginado), POST, PUT/{id}, DELETE/{id} (/clients)
"""
from __future__ import annotations

import io
from datetime import date
from typing import Any, Literal, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from passlib.hash import bcrypt
from pydantic import BaseModel, EmailStr, Field

from core.auth import CurrentUser, current_user, require_admin
from core.db import get_conn

router = APIRouter(prefix="/api/admin", tags=["admin-maestros"])

# R7-F4: secciones extraídas a archivos por entidad para reducir el tamaño
# de este god-module (era 2.5k líneas). Las URLs públicas son idénticas
# (todos comparten el prefix /api/admin del router padre).
from routers.mantenedores_empresas import router as _empresas_router  # noqa: E402
from routers.mantenedores_users import router as _users_router  # noqa: E402
from routers.mantenedores_drivers import router as _drivers_router  # noqa: E402
# Helpers compartidos que el resto de secciones de mantenedores.py todavía
# usa (vehicles, dotación, capacitaciones). Se re-exportan con los nombres
# privados originales para evitar romper código no migrado.
from routers.mantenedores_shared import (  # noqa: E402
    refresh_state_maestros as _refresh_state_maestros,
    require_fleet_access,
    enforce_fleet_empresa as _enforce_fleet_empresa,
)
from routers.mantenedores_drivers import (  # noqa: E402
    DriverIn, DriverUpdate, DriverOut,
    fetch_driver as _fetch_driver,
)

router.include_router(_empresas_router)
router.include_router(_users_router)
router.include_router(_drivers_router)


from routers.mantenedores_vehicles import router as _vehicles_router  # noqa: E402
from routers.mantenedores_vehicles import (  # noqa: E402
    VehicleIn, VehicleUpdate, VehicleOut,
    fetch_vehicle as _fetch_vehicle,
)
router.include_router(_vehicles_router)


# ============================================================================
# Dotación diaria — extraída en R7-F4 a mantenedores_dotacion.py
# ============================================================================
from routers.mantenedores_dotacion import router as _dotacion_router  # noqa: E402
from routers.mantenedores_dotacion import (  # noqa: E402
    DotacionEstado, DotacionUpdate, DotacionRowOut,
)
# Helper privado todavía usado por algunos uploads bulk:
from routers.mantenedores_shared import can_access_empresa as _can_access_empresa  # noqa: E402
router.include_router(_dotacion_router)


# ============================================================================
# Clients — extraídos en R7-F4 a mantenedores_clients.py
# ============================================================================
from routers.mantenedores_clients import router as _clients_router  # noqa: E402
from routers.mantenedores_clients import (  # noqa: E402
    ClientIn, ClientUpdate, ClientOut, ClientsPage,
)
router.include_router(_clients_router)


# ============================================================================
# Excel template + upload masivo (drivers, vehicles, dotacion)
# ============================================================================
class BulkUploadResult(BaseModel):
    created: int = 0
    updated: int = 0
    errors: list[str] = []


def _xlsx_response(filename: str, headers: list[str], rows: list[list[Any]]) -> StreamingResponse:
    """Genera un xlsx in-memory con headers + rows. Usa openpyxl (ya en deps)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    # Formatear header en negrita
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _read_xlsx_rows(file_bytes: bytes) -> tuple[list[str], list[list[Any]]]:
    """Devuelve (headers, rows). Headers normalizados a lower."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip().lower() if h is not None else "" for h in headers_raw]
    rows = [list(r) for r in rows_iter if any(c is not None for c in r)]
    return headers, rows


# ----- Drivers -----
@router.get("/drivers/template")
def drivers_template(empresa_id: int = Query(...),
                      user: CurrentUser = Depends(require_fleet_access)) -> StreamingResponse:
    _enforce_fleet_empresa(user, empresa_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            """SELECT driver_id, name, phone, license, vehicle_id, vehicle_name, active
                FROM fpoc.drivers WHERE empresa_id = ? ORDER BY driver_id""",
            empresa_id,
        )
        rows = [
            [r.driver_id, r.name, r.phone or "", r.license or "",
             int(r.vehicle_id), r.vehicle_name or "", 1 if r.active else 0]
            for r in cur.fetchall()
        ]
    headers = ["driver_id", "name", "phone", "license", "vehicle_id", "vehicle_name", "active"]
    return _xlsx_response(f"drivers_empresa_{empresa_id}.xlsx", headers, rows)


@router.post("/drivers/upload", response_model=BulkUploadResult)
async def drivers_upload(
    empresa_id: int = Query(...),
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_fleet_access),
) -> BulkUploadResult:
    _enforce_fleet_empresa(user, empresa_id)
    headers, rows = _read_xlsx_rows(await file.read())
    # phone es obligatorio al CREAR un driver (contacto WhatsApp).
    # En update se mantiene opcional (no se borra si no viene).
    required = ["driver_id", "name", "phone", "vehicle_id", "vehicle_name"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(400, f"Faltan columnas requeridas: {missing}")
    idx = {h: i for i, h in enumerate(headers)}
    result = BulkUploadResult()
    with get_conn() as cn:
        cur = cn.cursor()
        for line_num, r in enumerate(rows, start=2):  # row 1 es header
            try:
                driver_id = str(r[idx["driver_id"]]).strip()
                if not driver_id:
                    continue
                name = str(r[idx["name"]] or "").strip()
                phone = str(r[idx["phone"]] or "").strip() or None
                lic = str(r[idx["license"]] or "").strip() or None if "license" in idx else None
                vehicle_id = int(r[idx["vehicle_id"]])
                vehicle_name = str(r[idx["vehicle_name"]] or "").strip()
                active = bool(int(r[idx["active"]] or 1)) if "active" in idx else True
                # Existe? -> update, sino insert (con validación phone obligatorio)
                cur.execute("SELECT 1 FROM fpoc.drivers WHERE driver_id = ?", driver_id)
                existing = cur.fetchone()
                if not existing and not phone:
                    raise ValueError("phone obligatorio para nuevo driver")
                if existing:
                    cur.execute(
                        """UPDATE fpoc.drivers SET name=?, phone=?, license=?, empresa_id=?,
                                  vehicle_id=?, vehicle_name=?, active=?, updated_at=CURRENT_TIMESTAMP
                                WHERE driver_id=?""",
                        name, phone, lic, empresa_id, vehicle_id, vehicle_name,
                        1 if active else 0, driver_id,
                    )
                    result.updated += 1
                else:
                    cur.execute(
                        """INSERT INTO fpoc.drivers
                            (driver_id, name, phone, license, empresa_id, vehicle_id, vehicle_name, active)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                        driver_id, name, phone, lic, empresa_id, vehicle_id, vehicle_name,
                        1 if active else 0,
                    )
                    result.created += 1
            except Exception as e:  # noqa: BLE001
                result.errors.append(f"fila {line_num}: {type(e).__name__}: {e}")
        cn.commit()
    _refresh_state_maestros()
    return result


# ----- Vehicles -----
@router.get("/vehicles/template")
def vehicles_template(empresa_id: int = Query(...),
                       user: CurrentUser = Depends(require_fleet_access)) -> StreamingResponse:
    _enforce_fleet_empresa(user, empresa_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            """SELECT vehicle_id, name, type, plate, capacity_m3, driver_id, driver_name,
                       depot_lat, depot_lon, year, active
                FROM fpoc.vehicles WHERE empresa_id = ? ORDER BY vehicle_id""",
            empresa_id,
        )
        rows = [
            [int(r.vehicle_id), r.name, r.type or "", r.plate or "",
             int(r.capacity_m3 or 0), r.driver_id or "", r.driver_name or "",
             float(r.depot_lat), float(r.depot_lon),
             int(r.year) if r.year is not None else "",
             1 if r.active else 0]
            for r in cur.fetchall()
        ]
    headers = ["vehicle_id", "name", "type", "plate", "capacity_m3",
               "driver_id", "driver_name", "depot_lat", "depot_lon", "year", "active"]
    return _xlsx_response(f"vehicles_empresa_{empresa_id}.xlsx", headers, rows)


@router.post("/vehicles/upload", response_model=BulkUploadResult)
async def vehicles_upload(
    empresa_id: int = Query(...),
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_fleet_access),
) -> BulkUploadResult:
    _enforce_fleet_empresa(user, empresa_id)
    headers, rows = _read_xlsx_rows(await file.read())
    required = ["vehicle_id", "name", "type", "plate", "capacity_m3"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(400, f"Faltan columnas requeridas: {missing}")
    idx = {h: i for i, h in enumerate(headers)}
    result = BulkUploadResult()
    with get_conn() as cn:
        cur = cn.cursor()
        for line_num, r in enumerate(rows, start=2):
            try:
                vid = int(r[idx["vehicle_id"]])
                name = str(r[idx["name"]] or "").strip()
                vtype = str(r[idx["type"]] or "").strip()
                plate = str(r[idx["plate"]] or "").strip()
                cap = int(r[idx["capacity_m3"]] or 0)
                drv_id = str(r[idx["driver_id"]] or "").strip() or None if "driver_id" in idx else None
                drv_name = str(r[idx["driver_name"]] or "").strip() or None if "driver_name" in idx else None
                lat = float(r[idx["depot_lat"]] or -33.45) if "depot_lat" in idx else -33.45
                lon = float(r[idx["depot_lon"]] or -70.66) if "depot_lon" in idx else -70.66
                year = int(r[idx["year"]]) if "year" in idx and r[idx["year"]] not in (None, "") else None
                active = bool(int(r[idx["active"]] or 1)) if "active" in idx else True
                cur.execute("SELECT 1 FROM fpoc.vehicles WHERE vehicle_id = ?", vid)
                if cur.fetchone():
                    cur.execute(
                        """UPDATE fpoc.vehicles SET empresa_id=?, name=?, type=?, plate=?,
                                  capacity_m3=?, driver_id=?, driver_name=?, depot_lat=?,
                                  depot_lon=?, year=?, active=?, updated_at=CURRENT_TIMESTAMP
                                WHERE vehicle_id=?""",
                        empresa_id, name, vtype, plate, cap, drv_id, drv_name, lat, lon,
                        year, 1 if active else 0, vid,
                    )
                    result.updated += 1
                else:
                    cur.execute(
                        """INSERT INTO fpoc.vehicles
                            (vehicle_id, empresa_id, name, type, plate, capacity_m3, driver_id,
                             driver_name, depot_lat, depot_lon, year, active)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        vid, empresa_id, name, vtype, plate, cap, drv_id, drv_name, lat, lon,
                        year, 1 if active else 0,
                    )
                    result.created += 1
            except Exception as e:  # noqa: BLE001
                result.errors.append(f"fila {line_num}: {type(e).__name__}: {e}")
        cn.commit()
    _refresh_state_maestros()
    return result


# ----- Dotación -----
@router.get("/dotacion-diaria/template")
def dotacion_template(
    fecha: date = Query(...),
    empresa_id: Optional[int] = Query(default=None),
    user: CurrentUser = Depends(current_user),
) -> StreamingResponse:
    """Template pre-rellenado con drivers de la(s) empresa(s) y su estado actual."""
    empresa_ids = _dotacion_empresa_ids(user, empresa_id)
    for eid in empresa_ids:
        _can_access_empresa(user, eid)
    rows_data = _fetch_dotacion_rows(fecha, empresa_ids)
    headers = ["empresa_id", "driver_id", "driver_name", "vehicle_id", "estado", "motivo"]
    rows = [
        [r.empresa_id, r.driver_id or "", r.driver_name or "",
         r.vehicle_id if r.vehicle_id is not None else "",
         r.estado, r.motivo or ""]
        for r in rows_data
    ]
    suffix = f"empresa_{empresa_id}" if empresa_id else "todas"
    return _xlsx_response(f"dotacion_{fecha.isoformat()}_{suffix}.xlsx", headers, rows)


@router.post("/dotacion-diaria/upload", response_model=BulkUploadResult)
async def dotacion_upload(
    fecha: date = Query(...),
    file: UploadFile = File(...),
    user: CurrentUser = Depends(current_user),
) -> BulkUploadResult:
    headers, rows = _read_xlsx_rows(await file.read())
    required = ["empresa_id", "driver_id", "estado"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(400, f"Faltan columnas requeridas: {missing}")
    idx = {h: i for i, h in enumerate(headers)}
    valid_estados = {"disponible", "ausente", "licencia", "mantencion", "baja", "reemplazo"}
    result = BulkUploadResult()
    with get_conn() as cn:
        cur = cn.cursor()
        for line_num, r in enumerate(rows, start=2):
            try:
                eid = int(r[idx["empresa_id"]])
                _can_access_empresa(user, eid)
                drv_id = str(r[idx["driver_id"]]).strip() or None
                vid = (int(r[idx["vehicle_id"]])
                        if "vehicle_id" in idx and r[idx["vehicle_id"]] not in (None, "") else None)
                estado = str(r[idx["estado"]] or "disponible").strip().lower()
                if estado not in valid_estados:
                    raise ValueError(f"estado inválido: {estado}")
                motivo = str(r[idx["motivo"]] or "").strip() or None if "motivo" in idx else None
                if drv_id is None and vid is None:
                    raise ValueError("driver_id o vehicle_id requerido")
                # Buscar override existente
                if drv_id:
                    cur.execute(
                        "SELECT dotacion_id FROM fpoc.dotacion_diaria WHERE fecha=? AND empresa_id=? AND driver_id=?",
                        fecha.isoformat(), eid, drv_id,
                    )
                else:
                    cur.execute(
                        "SELECT dotacion_id FROM fpoc.dotacion_diaria WHERE fecha=? AND empresa_id=? AND vehicle_id=?",
                        fecha.isoformat(), eid, vid,
                    )
                existing = cur.fetchone()
                if existing:
                    cur.execute(
                        """UPDATE fpoc.dotacion_diaria
                                SET vehicle_id=?, estado=?, motivo=?, updated_by_user_id=?, updated_at=CURRENT_TIMESTAMP
                              WHERE dotacion_id=?""",
                        vid, estado, motivo, user.user_id, int(existing.dotacion_id),
                    )
                    result.updated += 1
                else:
                    cur.execute(
                        """INSERT INTO fpoc.dotacion_diaria
                            (fecha, empresa_id, driver_id, vehicle_id, estado, motivo, updated_by_user_id)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        fecha.isoformat(), eid, drv_id, vid, estado, motivo, user.user_id,
                    )
                    result.created += 1
            except HTTPException:
                raise
            except Exception as e:  # noqa: BLE001
                result.errors.append(f"fila {line_num}: {type(e).__name__}: {e}")
        cn.commit()
    return result


# ============================================================================
# Driver documents — extraído en R7-F4 a mantenedores_documents_driver.py
# ============================================================================
from routers.mantenedores_documents_driver import router as _doc_driver_router  # noqa: E402
from routers.mantenedores_documents_driver import DriverDocOut  # noqa: E402
router.include_router(_doc_driver_router)


# ============================================================================
# Capacitaciones — extraído en R7-F4 a mantenedores_capacitaciones.py
# ============================================================================
from routers.mantenedores_capacitaciones import router as _capacitaciones_router  # noqa: E402
from routers.mantenedores_capacitaciones import (  # noqa: E402
    CapacitacionModuloIn, CapacitacionModuloUpdate, CapacitacionModuloOut,
    DriverCapacitacionIn, DriverCapacitacionUpdate, DriverCapacitacionOut,
)
router.include_router(_capacitaciones_router)


class WhatsAppInviteRequest(BaseModel):
    phone_e164: str = Field(min_length=8, max_length=20)
    name: Optional[str] = Field(default=None, max_length=200)
    role_hint: Optional[str] = Field(
        default=None,
        pattern="^(driver|manager|contacto)$",
        description="Pista del rol para el mensaje de bienvenida",
    )
    # Opcionales para registrar opt-in en la entidad correcta una vez confirmen:
    driver_id: Optional[str] = Field(default=None, max_length=20)
    user_id: Optional[int] = Field(default=None, ge=1)
    contact_id: Optional[int] = Field(default=None, ge=1)
    custom_message: Optional[str] = Field(default=None, max_length=500)


class WhatsAppInviteResponse(BaseModel):
    ok: bool
    status: str         # 'sent' | 'dry_run' | 'error'
    twilio_sid: Optional[str] = None
    error: Optional[str] = None
    sandbox_warning: Optional[str] = None
    target_phone: str
    body_preview: str


@router.post("/whatsapp/invite", response_model=WhatsAppInviteResponse)
def whatsapp_invite(
    req: WhatsAppInviteRequest,
    user: CurrentUser = Depends(current_user),
) -> WhatsAppInviteResponse:
    """Envía un mensaje proactivo de WhatsApp invitando al usuario a opt-in.

    Permisos:
    - admin/ops: cualquier número.
    - transport_manager: solo números de drivers/contactos de SU empresa.
    """
    from routers.notifications import send_whatsapp, TwilioConfig
    if not user.is_falabella and user.role != "transport_manager":
        raise HTTPException(403, "Sin permisos para invitar")

    phone = req.phone_e164.strip()
    if not phone.startswith("+"):
        raise HTTPException(400, "phone_e164 debe empezar con + (formato E.164)")

    # Scope para transport_manager: validar que phone pertenezca a SU empresa
    if user.role == "transport_manager":
        with get_conn() as cn:
            cur = cn.cursor()
            # Driver de su empresa?
            cur.execute(
                "SELECT 1 FROM fpoc.drivers WHERE phone = ? AND empresa_id = ?",
                phone, user.empresa_id,
            )
            ok_driver = cur.fetchone() is not None
            # Contacto de su empresa?
            cur.execute(
                "SELECT 1 FROM fpoc.empresa_contactos WHERE phone_e164 = ? AND empresa_id = ?",
                phone, user.empresa_id,
            )
            ok_contact = cur.fetchone() is not None
            if not (ok_driver or ok_contact):
                raise HTTPException(403, "Ese número no pertenece a tu empresa")

    nombre = req.name or "Usuario"
    role = req.role_hint or "contacto"
    role_label = {
        "driver": "conductor",
        "manager": "responsable de transporte",
        "contacto": "contacto",
    }.get(role, "contacto")

    if req.custom_message:
        body = req.custom_message
    else:
        body = (
            f"Hola {nombre}, te están invitando como *{role_label}* a la Torre de "
            f"Control ValueData × Falabella.\n\n"
            f"Para activar las alertas y poder reportar entregas por WhatsApp, "
            f"respondé *SI* a este mensaje.\n\n"
            f"Si fue un error, ignorá y te damos de baja."
        )

    # Template Meta-approved vd_invitacion (1 var = nombre).
    # Fallback freeform si el template falla — preserva body custom_message si
    # el usuario lo proveyó.
    from core.twilio_templates import invitacion_sid
    content_sid = invitacion_sid()
    res = None
    # Si el usuario forzó un custom_message, NO podemos usar el template
    # (template tiene texto fijo). En ese caso, mantener freeform.
    use_template = bool(content_sid) and not req.custom_message
    if use_template:
        try:
            from routers.comments import _sanitize_template_var as _sanvar
            res = send_whatsapp(
                content_sid=content_sid,
                content_variables={"1": _sanvar(nombre) or "Usuario"},
                targets=[(req.user_id, phone)],
                subject="Invitación WhatsApp",
                triggered_by="invite",
            )
        except Exception as e:  # noqa: BLE001
            from loguru import logger as _log
            _log.warning(f"[mantenedores] template vd_invitacion falló, fallback freeform: {e}")
            res = None
    if res is None:
        res = send_whatsapp(
            body=body,
            targets=[(req.user_id, phone)],
            subject="Invitación WhatsApp",
            triggered_by="invite",
        )

    cfg = TwilioConfig.from_env()
    sandbox_warning = None
    if "+14155238886" in cfg.from_number:
        sandbox_warning = (
            "Estás usando el sandbox de Twilio (+14155238886). El destinatario tiene "
            "que haber enviado 'join <code>' desde su WhatsApp para poder recibir "
            "mensajes. Configurá TWILIO_WHATSAPP_FROM con tu número propio para evitar esto."
        )

    if res.results:
        r = res.results[0]
        return WhatsAppInviteResponse(
            ok=(r.status in ("sent", "dry_run")),
            status=r.status,
            twilio_sid=r.twilio_sid,
            error=r.error,
            sandbox_warning=sandbox_warning,
            target_phone=phone,
            body_preview=body[:200],
        )
    # Sin results = notifications deshabilitado
    return WhatsAppInviteResponse(
        ok=False,
        status="disabled",
        error="NOTIFICATIONS_ENABLED=false en el backend",
        sandbox_warning=sandbox_warning,
        target_phone=phone,
        body_preview=body[:200],
    )


# validate/unvalidate de capacitaciones movidos a mantenedores_capacitaciones.py (R7-F4)


# ============================================================================
# Document types — extraído en R7-F4 a mantenedores_doctypes.py
# ============================================================================
from routers.mantenedores_doctypes import router as _doctypes_router  # noqa: E402
from routers.mantenedores_doctypes import (  # noqa: E402
    DocTypeIn, DocTypeUpdate, DocTypeOut,
)
router.include_router(_doctypes_router)


# ============================================================================
# Entity documents (empresas + vehicles) — extraídos en R7-F4
# a mantenedores_documents_entity.py
# ============================================================================
from routers.mantenedores_documents_entity import router as _doc_entity_router  # noqa: E402
from routers.mantenedores_documents_entity import EntityDocOut  # noqa: E402
router.include_router(_doc_entity_router)
