"""Mantenedor de maestros: CRUD admin sobre empresas, users, drivers, vehicles, clients.

Todos los endpoints requieren rol `falabella_admin`. Tras cada mutación de
drivers/vehicles/clients se llama `STATE.reload_maestros()` para que el cache
in-memory quede consistente con la DB.

Endpoints (prefijo /api/admin):
  Empresas:  GET, POST, PUT/{id}, DELETE/{id}     (/empresas)
  Users:     GET, POST, PUT/{id}, DELETE/{id}, POST/{id}/reset-password (/users)
  Drivers:   GET, POST, PUT/{id}, DELETE/{id}     (/drivers)
  Vehicles:  GET, POST, PUT/{id}, DELETE/{id}     (/vehicles)
  Clients:   GET (paginado), POST, PUT/{id}, DELETE/{id} (/clients)
"""
from __future__ import annotations

import io
from datetime import date
from typing import Any, Literal, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from passlib.hash import bcrypt
from pydantic import BaseModel, EmailStr, Field

from auth import CurrentUser, current_user, require_admin
from db import get_conn

router = APIRouter(prefix="/api/admin", tags=["admin-maestros"])


# ============================================================================
# Empresas
# ============================================================================
class EmpresaIn(BaseModel):
    empresa_id: int = Field(ge=1)
    nombre: str = Field(min_length=1, max_length=100)
    activo: bool = True


class EmpresaUpdate(BaseModel):
    nombre: Optional[str] = Field(default=None, min_length=1, max_length=100)
    activo: Optional[bool] = None


class EmpresaOut(BaseModel):
    empresa_id: int
    nombre: str
    activo: bool
    created_at: Optional[str] = None


def _empresa_row(r) -> EmpresaOut:
    created = r.created_at
    return EmpresaOut(
        empresa_id=int(r.empresa_id),
        nombre=r.nombre,
        activo=bool(r.activo),
        created_at=created.isoformat() if hasattr(created, "isoformat") else (created or None),
    )


@router.get("/empresas", response_model=list[EmpresaOut])
def list_empresas(_: CurrentUser = Depends(require_admin)) -> list[EmpresaOut]:
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            "SELECT empresa_id, nombre, activo, created_at FROM fpoc.empresas_transporte ORDER BY empresa_id"
        )
        return [_empresa_row(r) for r in cur.fetchall()]


@router.post("/empresas", response_model=EmpresaOut)
def create_empresa(req: EmpresaIn, _: CurrentUser = Depends(require_admin)) -> EmpresaOut:
    with get_conn() as cn:
        cur = cn.cursor()
        try:
            cur.execute(
                "INSERT INTO fpoc.empresas_transporte (empresa_id, nombre, activo) VALUES (?, ?, ?)",
                req.empresa_id, req.nombre, 1 if req.activo else 0,
            )
            cn.commit()
        except Exception as e:  # noqa: BLE001
            cn.rollback()
            raise HTTPException(409, f"empresa duplicada o inválida: {e}")
        cur.execute(
            "SELECT empresa_id, nombre, activo, created_at FROM fpoc.empresas_transporte WHERE empresa_id = ?",
            req.empresa_id,
        )
        return _empresa_row(cur.fetchone())


@router.put("/empresas/{empresa_id}", response_model=EmpresaOut)
def update_empresa(empresa_id: int, req: EmpresaUpdate,
                    _: CurrentUser = Depends(require_admin)) -> EmpresaOut:
    sets, params = [], []
    if req.nombre is not None:
        sets.append("nombre = ?"); params.append(req.nombre)
    if req.activo is not None:
        sets.append("activo = ?"); params.append(1 if req.activo else 0)
    if not sets:
        raise HTTPException(400, "nada que actualizar")
    params.append(empresa_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            f"UPDATE fpoc.empresas_transporte SET {', '.join(sets)} WHERE empresa_id = ?",
            *params,
        )
        if cur.rowcount == 0:
            raise HTTPException(404, "empresa no encontrada")
        cn.commit()
        cur.execute(
            "SELECT empresa_id, nombre, activo, created_at FROM fpoc.empresas_transporte WHERE empresa_id = ?",
            empresa_id,
        )
        return _empresa_row(cur.fetchone())


@router.delete("/empresas/{empresa_id}")
def delete_empresa(empresa_id: int, _: CurrentUser = Depends(require_admin)) -> dict:
    with get_conn() as cn:
        cur = cn.cursor()
        # Validar que no tenga usuarios asociados
        cur.execute("SELECT COUNT(*) FROM fpoc.users WHERE empresa_id = ?", empresa_id)
        n_users = int(cur.fetchone()[0])
        if n_users > 0:
            raise HTTPException(409, f"empresa tiene {n_users} usuarios; desactivar en vez de eliminar")
        cur.execute("DELETE FROM fpoc.empresas_transporte WHERE empresa_id = ?", empresa_id)
        if cur.rowcount == 0:
            raise HTTPException(404, "empresa no encontrada")
        cn.commit()
    return {"deleted": empresa_id}


# ============================================================================
# Users
# ============================================================================
ALLOWED_ROLES = {"falabella_admin", "falabella_ops", "transport_manager"}


class UserIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=4, max_length=128)
    display_name: str = Field(min_length=1, max_length=200)
    role: str = Field(pattern="^(falabella_admin|falabella_ops|transport_manager|driver)$")
    empresa_id: Optional[int] = None
    driver_id: Optional[str] = Field(default=None, max_length=20)
    activo: bool = True
    phone_e164: Optional[str] = Field(default=None, max_length=20)
    notify_whatsapp: bool = False


class UserUpdate(BaseModel):
    email: Optional[EmailStr] = None
    display_name: Optional[str] = Field(default=None, min_length=1, max_length=200)
    role: Optional[str] = Field(default=None, pattern="^(falabella_admin|falabella_ops|transport_manager|driver)$")
    empresa_id: Optional[int] = None
    driver_id: Optional[str] = Field(default=None, max_length=20)
    activo: Optional[bool] = None
    phone_e164: Optional[str] = Field(default=None, max_length=20)
    notify_whatsapp: Optional[bool] = None


class PasswordReset(BaseModel):
    new_password: str = Field(min_length=4, max_length=128)


class UserOut(BaseModel):
    user_id: int
    email: str
    display_name: str
    role: str
    empresa_id: Optional[int] = None
    empresa_nombre: Optional[str] = None
    driver_id: Optional[str] = None
    driver_name: Optional[str] = None
    activo: bool
    phone_e164: Optional[str] = None
    notify_whatsapp: bool
    created_at: Optional[str] = None
    last_login: Optional[str] = None


def _user_row(r) -> UserOut:
    def _iso(v):
        if v is None: return None
        return v.isoformat() if hasattr(v, "isoformat") else str(v)
    return UserOut(
        user_id=int(r.user_id), email=r.email, display_name=r.display_name,
        role=r.role,
        empresa_id=int(r.empresa_id) if r.empresa_id is not None else None,
        empresa_nombre=r.empresa_nombre,
        driver_id=str(r.driver_id) if r.driver_id is not None else None,
        driver_name=getattr(r, "driver_name", None),
        activo=bool(r.activo),
        phone_e164=r.phone_e164,
        notify_whatsapp=bool(r.notify_whatsapp),
        created_at=_iso(r.created_at),
        last_login=_iso(r.last_login),
    )


_USER_SELECT = """
    SELECT u.user_id, u.email, u.display_name, u.role, u.empresa_id,
           u.driver_id, u.activo, u.phone_e164, u.notify_whatsapp,
           u.created_at, u.last_login,
           e.nombre AS empresa_nombre,
           d.name AS driver_name
    FROM fpoc.users u
    LEFT JOIN fpoc.empresas_transporte e ON u.empresa_id = e.empresa_id
    LEFT JOIN fpoc.drivers d ON u.driver_id = d.driver_id
"""


@router.get("/users", response_model=list[UserOut])
def list_users(_: CurrentUser = Depends(require_admin)) -> list[UserOut]:
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(_USER_SELECT + " ORDER BY u.user_id")
        return [_user_row(r) for r in cur.fetchall()]


def _can_manage_user(actor: CurrentUser, target_role: str, target_empresa_id: Optional[int]) -> bool:
    """admin: todos. transport_manager: solo crear/editar drivers de su empresa."""
    if actor.is_admin:
        return True
    if actor.role == "transport_manager" and target_role == "driver":
        return target_empresa_id == actor.empresa_id
    return False


@router.post("/users", response_model=UserOut)
def create_user(req: UserIn, user: CurrentUser = Depends(current_user)) -> UserOut:
    if not _can_manage_user(user, req.role, req.empresa_id):
        raise HTTPException(403, "Sin permisos para crear este usuario")
    if req.role == "transport_manager" and req.empresa_id is None:
        raise HTTPException(400, "transport_manager requiere empresa_id")
    if req.role == "driver":
        if req.driver_id is None or req.empresa_id is None:
            raise HTTPException(400, "rol driver requiere driver_id y empresa_id")
        # Validar que el driver existe y es de esa empresa
        with get_conn() as cn:
            cur = cn.cursor()
            cur.execute("SELECT empresa_id FROM fpoc.drivers WHERE driver_id = ?", req.driver_id)
            row = cur.fetchone()
            if not row:
                raise HTTPException(400, f"driver_id {req.driver_id} no existe")
            if int(row.empresa_id) != req.empresa_id:
                raise HTTPException(400, "driver_id no pertenece a la empresa indicada")
    with get_conn() as cn:
        cur = cn.cursor()
        try:
            cur.execute(
                """
                INSERT INTO fpoc.users
                  (email, password_hash, display_name, role, empresa_id, driver_id, activo,
                   phone_e164, notify_whatsapp)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                RETURNING user_id
                """,
                req.email.lower(), bcrypt.hash(req.password), req.display_name,
                req.role, req.empresa_id, req.driver_id, 1 if req.activo else 0,
                req.phone_e164, 1 if req.notify_whatsapp else 0,
            )
            new_id = int(cur.fetchone()[0])
            cn.commit()
        except Exception as e:  # noqa: BLE001
            cn.rollback()
            raise HTTPException(409, f"email duplicado o datos inválidos: {e}")
        cur.execute(_USER_SELECT + " WHERE u.user_id = ?", new_id)
        return _user_row(cur.fetchone())


@router.put("/users/{user_id}", response_model=UserOut)
def update_user(user_id: int, req: UserUpdate,
                 _: CurrentUser = Depends(require_admin)) -> UserOut:
    sets, params = [], []
    for field, col in [
        ("email", "email"), ("display_name", "display_name"),
        ("role", "role"), ("empresa_id", "empresa_id"),
        ("phone_e164", "phone_e164"),
    ]:
        v = getattr(req, field)
        if v is not None:
            sets.append(f"{col} = ?")
            params.append(v.lower() if field == "email" else v)
    if req.activo is not None:
        sets.append("activo = ?"); params.append(1 if req.activo else 0)
    if req.notify_whatsapp is not None:
        sets.append("notify_whatsapp = ?"); params.append(1 if req.notify_whatsapp else 0)
    if not sets:
        raise HTTPException(400, "nada que actualizar")
    params.append(user_id)
    with get_conn() as cn:
        cur = cn.cursor()
        try:
            cur.execute(f"UPDATE fpoc.users SET {', '.join(sets)} WHERE user_id = ?", *params)
        except Exception as e:  # noqa: BLE001
            cn.rollback()
            raise HTTPException(409, f"datos inválidos: {e}")
        if cur.rowcount == 0:
            raise HTTPException(404, "user no encontrado")
        cn.commit()
        cur.execute(_USER_SELECT + " WHERE u.user_id = ?", user_id)
        return _user_row(cur.fetchone())


@router.post("/users/{user_id}/reset-password")
def reset_password(user_id: int, req: PasswordReset,
                    _: CurrentUser = Depends(require_admin)) -> dict:
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            "UPDATE fpoc.users SET password_hash = ? WHERE user_id = ?",
            bcrypt.hash(req.new_password), user_id,
        )
        if cur.rowcount == 0:
            raise HTTPException(404, "user no encontrado")
        cn.commit()
    return {"reset": user_id}


@router.delete("/users/{user_id}")
def delete_user(user_id: int, admin: CurrentUser = Depends(require_admin)) -> dict:
    if user_id == admin.user_id:
        raise HTTPException(400, "no puedes eliminarte a ti mismo")
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute("DELETE FROM fpoc.users WHERE user_id = ?", user_id)
        if cur.rowcount == 0:
            raise HTTPException(404, "user no encontrado")
        cn.commit()
    return {"deleted": user_id}


# ============================================================================
# Drivers
# ============================================================================
class DriverIn(BaseModel):
    driver_id: str = Field(min_length=1, max_length=20)
    name: str = Field(min_length=1, max_length=200)
    phone: Optional[str] = Field(default=None, max_length=50)
    license: Optional[str] = Field(default="A-3 Profesional", max_length=50)
    empresa_id: int = Field(ge=1)
    vehicle_id: int = Field(ge=1)
    vehicle_name: str = Field(min_length=1, max_length=50)
    rating: float = Field(default=4.5, ge=0.0, le=5.0)
    deliveries_30d: int = Field(default=0, ge=0)
    fail_rate_30d: float = Field(default=0.10, ge=0.0, le=1.0)
    joined_at: Optional[date] = None
    active: bool = True


class DriverUpdate(BaseModel):
    name: Optional[str] = Field(default=None, min_length=1, max_length=200)
    phone: Optional[str] = Field(default=None, max_length=50)
    license: Optional[str] = Field(default=None, max_length=50)
    empresa_id: Optional[int] = Field(default=None, ge=1)
    vehicle_id: Optional[int] = Field(default=None, ge=1)
    vehicle_name: Optional[str] = Field(default=None, min_length=1, max_length=50)
    rating: Optional[float] = Field(default=None, ge=0.0, le=5.0)
    deliveries_30d: Optional[int] = Field(default=None, ge=0)
    fail_rate_30d: Optional[float] = Field(default=None, ge=0.0, le=1.0)
    joined_at: Optional[date] = None
    active: Optional[bool] = None


class DriverOut(BaseModel):
    driver_id: str
    name: str
    phone: Optional[str] = None
    license: Optional[str] = None
    empresa_id: Optional[int] = None
    empresa_nombre: Optional[str] = None
    vehicle_id: int
    vehicle_name: str
    rating: float
    deliveries_30d: int
    fail_rate_30d: float
    joined_at: Optional[str] = None
    active: bool
    is_problem_hidden: bool = False


def _driver_row(r) -> DriverOut:
    joined = r.joined_at
    return DriverOut(
        driver_id=r.driver_id, name=r.name, phone=r.phone, license=r.license,
        empresa_id=int(r.empresa_id) if r.empresa_id is not None else None,
        empresa_nombre=getattr(r, "empresa_nombre", None),
        vehicle_id=int(r.vehicle_id), vehicle_name=r.vehicle_name,
        rating=float(r.rating), deliveries_30d=int(r.deliveries_30d),
        fail_rate_30d=float(r.fail_rate_30d),
        joined_at=joined.isoformat() if hasattr(joined, "isoformat") else (joined or None),
        active=bool(r.active),
        is_problem_hidden=bool(r.is_problem_hidden),
    )


def _refresh_state_maestros() -> None:
    """Llamar tras CRUD de drivers/vehicles/clients."""
    try:
        from state import STATE
        STATE.reload_maestros()
    except Exception:  # noqa: BLE001
        pass  # Tolerante: el endpoint igual devolvió OK al cliente.


def require_fleet_access(user: CurrentUser = Depends(current_user)) -> CurrentUser:
    """Drivers/Vehicles: permite admin/ops o transport_manager (scopeado a su empresa).

    Los endpoints que la usan deben validar empresa_id contra user.empresa_id en el
    body/recurso vía _enforce_fleet_empresa.
    """
    if user.is_falabella:
        return user
    if user.role == "transport_manager" and user.empresa_id is not None:
        return user
    raise HTTPException(403, "Requiere rol falabella o transport_manager con empresa")


def _enforce_fleet_empresa(user: CurrentUser, empresa_id: Optional[int]) -> None:
    """transport_manager solo puede tocar recursos de SU empresa."""
    if user.is_falabella:
        return
    if empresa_id is None:
        raise HTTPException(400, "empresa_id requerido")
    if user.empresa_id != empresa_id:
        raise HTTPException(403, "Solo podés gestionar tu empresa")


def _fetch_driver(driver_id: str) -> DriverOut:
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            """SELECT d.driver_id, d.name, d.phone, d.license, d.empresa_id,
                       e.nombre AS empresa_nombre,
                       d.vehicle_id, d.vehicle_name,
                       d.rating, d.deliveries_30d, d.fail_rate_30d, d.joined_at, d.active,
                       d.is_problem_hidden
                FROM fpoc.drivers d
                LEFT JOIN fpoc.empresas_transporte e ON e.empresa_id = d.empresa_id
                WHERE d.driver_id = ?""",
            driver_id,
        )
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, "driver no encontrado")
        return _driver_row(r)


def _fetch_vehicle(vehicle_id: int) -> VehicleOut:
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            """SELECT v.vehicle_id, v.empresa_id, e.nombre AS empresa_nombre,
                       v.name, v.type, v.plate, v.capacity_m3, v.driver_id, v.driver_name,
                       v.depot_lat, v.depot_lon, v.year, v.active, v.is_problem_hidden
                FROM fpoc.vehicles v
                LEFT JOIN fpoc.empresas_transporte e ON e.empresa_id = v.empresa_id
                WHERE v.vehicle_id = ?""",
            vehicle_id,
        )
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, "vehicle no encontrado")
        return _vehicle_row(r)


@router.get("/drivers", response_model=list[DriverOut])
def list_drivers(user: CurrentUser = Depends(require_fleet_access)) -> list[DriverOut]:
    where = "" if user.is_falabella else "WHERE d.empresa_id = ?"
    params: list = [] if user.is_falabella else [user.empresa_id]
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            f"""SELECT d.driver_id, d.name, d.phone, d.license, d.empresa_id,
                       e.nombre AS empresa_nombre,
                       d.vehicle_id, d.vehicle_name,
                       d.rating, d.deliveries_30d, d.fail_rate_30d, d.joined_at, d.active,
                       d.is_problem_hidden
                FROM fpoc.drivers d
                LEFT JOIN fpoc.empresas_transporte e ON e.empresa_id = d.empresa_id
                {where}
                ORDER BY d.empresa_id, d.vehicle_id""",
            *params,
        )
        return [_driver_row(r) for r in cur.fetchall()]


@router.post("/drivers", response_model=DriverOut)
def create_driver(req: DriverIn, user: CurrentUser = Depends(require_fleet_access)) -> DriverOut:
    _enforce_fleet_empresa(user, req.empresa_id)
    with get_conn() as cn:
        cur = cn.cursor()
        try:
            cur.execute(
                """INSERT INTO fpoc.drivers
                    (driver_id, name, phone, license, empresa_id, vehicle_id, vehicle_name,
                     rating, deliveries_30d, fail_rate_30d, joined_at, active)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                req.driver_id, req.name, req.phone, req.license,
                req.empresa_id, req.vehicle_id, req.vehicle_name,
                req.rating, req.deliveries_30d, req.fail_rate_30d,
                req.joined_at.isoformat() if req.joined_at else None,
                1 if req.active else 0,
            )
            cn.commit()
        except Exception as e:  # noqa: BLE001
            cn.rollback()
            raise HTTPException(409, f"driver_id duplicado o datos inválidos: {e}")
    _refresh_state_maestros()
    return _fetch_driver(req.driver_id)


@router.put("/drivers/{driver_id}", response_model=DriverOut)
def update_driver(driver_id: str, req: DriverUpdate,
                   user: CurrentUser = Depends(require_fleet_access)) -> DriverOut:
    # Verificar que el driver es de la empresa del manager (si aplica)
    existing = _fetch_driver(driver_id)
    _enforce_fleet_empresa(user, existing.empresa_id)
    if req.empresa_id is not None:
        _enforce_fleet_empresa(user, req.empresa_id)
    sets, params = [], []
    for field in ["name", "phone", "license", "empresa_id", "vehicle_id", "vehicle_name",
                   "rating", "deliveries_30d", "fail_rate_30d"]:
        v = getattr(req, field)
        if v is not None:
            sets.append(f"{field} = ?"); params.append(v)
    if req.joined_at is not None:
        sets.append("joined_at = ?"); params.append(req.joined_at.isoformat())
    if req.active is not None:
        sets.append("active = ?"); params.append(1 if req.active else 0)
    sets.append("updated_at = CURRENT_TIMESTAMP")
    if not sets:
        raise HTTPException(400, "nada que actualizar")
    params.append(driver_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(f"UPDATE fpoc.drivers SET {', '.join(sets)} WHERE driver_id = ?", *params)
        if cur.rowcount == 0:
            raise HTTPException(404, "driver no encontrado")
        cn.commit()
    _refresh_state_maestros()
    return _fetch_driver(driver_id)


@router.delete("/drivers/{driver_id}")
def delete_driver(driver_id: str, user: CurrentUser = Depends(require_fleet_access)) -> dict:
    existing = _fetch_driver(driver_id)
    _enforce_fleet_empresa(user, existing.empresa_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute("DELETE FROM fpoc.drivers WHERE driver_id = ?", driver_id)
        if cur.rowcount == 0:
            raise HTTPException(404, "driver no encontrado")
        cn.commit()
    _refresh_state_maestros()
    return {"deleted": driver_id}


# ============================================================================
# Vehicles
# ============================================================================
class VehicleIn(BaseModel):
    vehicle_id: int = Field(ge=1)
    empresa_id: int = Field(ge=1)
    name: str = Field(min_length=1, max_length=50)
    type: str = Field(min_length=1, max_length=50)
    plate: str = Field(min_length=1, max_length=20)
    capacity_m3: int = Field(ge=0)
    driver_id: Optional[str] = Field(default=None, max_length=20)
    driver_name: Optional[str] = Field(default=None, max_length=200)
    depot_lat: float = -33.45
    depot_lon: float = -70.66
    year: Optional[int] = Field(default=None, ge=1990, le=2100)
    active: bool = True


class VehicleUpdate(BaseModel):
    empresa_id: Optional[int] = Field(default=None, ge=1)
    name: Optional[str] = Field(default=None, min_length=1, max_length=50)
    type: Optional[str] = Field(default=None, min_length=1, max_length=50)
    plate: Optional[str] = Field(default=None, min_length=1, max_length=20)
    capacity_m3: Optional[int] = Field(default=None, ge=0)
    driver_id: Optional[str] = Field(default=None, max_length=20)
    driver_name: Optional[str] = Field(default=None, max_length=200)
    year: Optional[int] = Field(default=None, ge=1990, le=2100)
    active: Optional[bool] = None


class VehicleOut(BaseModel):
    vehicle_id: int
    empresa_id: Optional[int] = None
    empresa_nombre: Optional[str] = None
    name: str
    type: str
    plate: str
    capacity_m3: int
    driver_id: Optional[str] = None
    driver_name: Optional[str] = None
    depot_lat: float
    depot_lon: float
    year: Optional[int] = None
    active: bool
    is_problem_hidden: bool = False


def _vehicle_row(r) -> VehicleOut:
    return VehicleOut(
        vehicle_id=int(r.vehicle_id),
        empresa_id=int(r.empresa_id) if r.empresa_id is not None else None,
        empresa_nombre=getattr(r, "empresa_nombre", None),
        name=r.name, type=r.type, plate=r.plate,
        capacity_m3=int(r.capacity_m3),
        driver_id=r.driver_id, driver_name=r.driver_name,
        depot_lat=float(r.depot_lat), depot_lon=float(r.depot_lon),
        year=int(r.year) if r.year is not None else None,
        active=bool(r.active),
        is_problem_hidden=bool(r.is_problem_hidden),
    )


@router.get("/vehicles", response_model=list[VehicleOut])
def list_vehicles(user: CurrentUser = Depends(require_fleet_access)) -> list[VehicleOut]:
    where = "" if user.is_falabella else "WHERE v.empresa_id = ?"
    params: list = [] if user.is_falabella else [user.empresa_id]
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            f"""SELECT v.vehicle_id, v.empresa_id, e.nombre AS empresa_nombre,
                       v.name, v.type, v.plate, v.capacity_m3, v.driver_id, v.driver_name,
                       v.depot_lat, v.depot_lon, v.year, v.active, v.is_problem_hidden
                FROM fpoc.vehicles v
                LEFT JOIN fpoc.empresas_transporte e ON e.empresa_id = v.empresa_id
                {where}
                ORDER BY v.empresa_id, v.vehicle_id""",
            *params,
        )
        return [_vehicle_row(r) for r in cur.fetchall()]


@router.post("/vehicles", response_model=VehicleOut)
def create_vehicle(req: VehicleIn, user: CurrentUser = Depends(require_fleet_access)) -> VehicleOut:
    _enforce_fleet_empresa(user, req.empresa_id)
    with get_conn() as cn:
        cur = cn.cursor()
        try:
            cur.execute(
                """INSERT INTO fpoc.vehicles
                    (vehicle_id, empresa_id, name, type, plate, capacity_m3, driver_id, driver_name,
                     depot_lat, depot_lon, year, active)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                req.vehicle_id, req.empresa_id, req.name, req.type, req.plate, req.capacity_m3,
                req.driver_id, req.driver_name, req.depot_lat, req.depot_lon,
                req.year, 1 if req.active else 0,
            )
            cn.commit()
        except Exception as e:  # noqa: BLE001
            cn.rollback()
            raise HTTPException(409, f"vehicle_id duplicado o datos inválidos: {e}")
    _refresh_state_maestros()
    return _fetch_vehicle(req.vehicle_id)


@router.put("/vehicles/{vehicle_id}", response_model=VehicleOut)
def update_vehicle(vehicle_id: int, req: VehicleUpdate,
                    user: CurrentUser = Depends(require_fleet_access)) -> VehicleOut:
    existing = _fetch_vehicle(vehicle_id)
    _enforce_fleet_empresa(user, existing.empresa_id)
    if req.empresa_id is not None:
        _enforce_fleet_empresa(user, req.empresa_id)
    sets, params = [], []
    for field in ["empresa_id", "name", "type", "plate", "capacity_m3", "driver_id", "driver_name", "year"]:
        v = getattr(req, field)
        if v is not None:
            sets.append(f"{field} = ?"); params.append(v)
    if req.active is not None:
        sets.append("active = ?"); params.append(1 if req.active else 0)
    sets.append("updated_at = CURRENT_TIMESTAMP")
    if not sets:
        raise HTTPException(400, "nada que actualizar")
    params.append(vehicle_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(f"UPDATE fpoc.vehicles SET {', '.join(sets)} WHERE vehicle_id = ?", *params)
        if cur.rowcount == 0:
            raise HTTPException(404, "vehicle no encontrado")
        cn.commit()
    _refresh_state_maestros()
    return _fetch_vehicle(vehicle_id)


@router.delete("/vehicles/{vehicle_id}")
def delete_vehicle(vehicle_id: int, user: CurrentUser = Depends(require_fleet_access)) -> dict:
    existing = _fetch_vehicle(vehicle_id)
    _enforce_fleet_empresa(user, existing.empresa_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute("DELETE FROM fpoc.vehicles WHERE vehicle_id = ?", vehicle_id)
        if cur.rowcount == 0:
            raise HTTPException(404, "vehicle no encontrado")
        cn.commit()
    _refresh_state_maestros()
    return {"deleted": vehicle_id}


# ============================================================================
# Dotacion diaria
# ============================================================================
DotacionEstado = Literal["disponible", "ausente", "licencia", "mantencion", "baja", "reemplazo"]


class DotacionUpdate(BaseModel):
    fecha: date
    empresa_id: int = Field(ge=1)
    driver_id: Optional[str] = Field(default=None, max_length=20)
    vehicle_id: Optional[int] = Field(default=None, ge=1)
    estado: DotacionEstado = "disponible"
    motivo: Optional[str] = Field(default=None, max_length=500)


class DotacionRowOut(BaseModel):
    fecha: str
    empresa_id: int
    empresa_nombre: Optional[str] = None
    driver_id: Optional[str] = None
    driver_name: Optional[str] = None
    driver_active: bool = True
    default_vehicle_id: Optional[int] = None
    vehicle_id: Optional[int] = None
    vehicle_name: Optional[str] = None
    plate: Optional[str] = None
    vehicle_active: bool = True
    estado: DotacionEstado = "disponible"
    motivo: Optional[str] = None
    updated_at: Optional[str] = None


def _can_access_empresa(user: CurrentUser, empresa_id: int) -> None:
    if user.is_falabella:
        return
    if user.role == "transport_manager" and user.empresa_id == empresa_id:
        return
    raise HTTPException(403, "sin permisos para esa empresa")


def _dotacion_empresa_ids(user: CurrentUser, empresa_id: Optional[int]) -> list[int]:
    if not user.is_falabella:
        if user.empresa_id is None:
            return []
        return [int(user.empresa_id)]
    if empresa_id is not None:
        return [int(empresa_id)]
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute("SELECT empresa_id FROM fpoc.empresas_transporte WHERE activo = 1 ORDER BY empresa_id")
        return [int(r.empresa_id) for r in cur.fetchall()]


def _fetch_dotacion_rows(fecha: date, empresa_ids: list[int]) -> list[DotacionRowOut]:
    if not empresa_ids:
        return []
    marks = ",".join(["?"] * len(empresa_ids))
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            f"""
            SELECT empresa_id, nombre
            FROM fpoc.empresas_transporte
            WHERE empresa_id IN ({marks})
            """,
            *empresa_ids,
        )
        empresas = {int(r.empresa_id): r.nombre for r in cur.fetchall()}

        cur.execute(
            f"""
            SELECT d.driver_id, d.name, d.empresa_id, d.vehicle_id, d.vehicle_name, d.active
            FROM fpoc.drivers d
            WHERE d.empresa_id IN ({marks}) AND d.active = 1
            ORDER BY d.empresa_id, d.vehicle_id, d.name
            """,
            *empresa_ids,
        )
        drivers = cur.fetchall()

        cur.execute(
            f"""
            SELECT v.vehicle_id, v.empresa_id, v.name, v.plate, v.driver_id, v.driver_name, v.active
            FROM fpoc.vehicles v
            WHERE v.empresa_id IN ({marks}) AND v.active = 1
            ORDER BY v.empresa_id, v.vehicle_id
            """,
            *empresa_ids,
        )
        vehicles = cur.fetchall()
        vehicle_by_id = {int(v.vehicle_id): v for v in vehicles}

        cur.execute(
            f"""
            SELECT dotacion_id, fecha, empresa_id, driver_id, vehicle_id, estado, motivo, updated_at
            FROM fpoc.dotacion_diaria
            WHERE fecha = ? AND empresa_id IN ({marks})
            """,
            fecha.isoformat(), *empresa_ids,
        )
        overrides = cur.fetchall()

    by_driver = {str(r.driver_id): r for r in overrides if r.driver_id is not None}
    by_vehicle = {int(r.vehicle_id): r for r in overrides if r.vehicle_id is not None}
    used_vehicle_ids: set[int] = set()
    out: list[DotacionRowOut] = []

    for d in drivers:
        default_vid = int(d.vehicle_id) if d.vehicle_id is not None else None
        ov = by_driver.get(str(d.driver_id))
        if ov is None and default_vid is not None:
            ov = by_vehicle.get(default_vid)
        vid = int(ov.vehicle_id) if ov is not None and ov.vehicle_id is not None else default_vid
        vehicle = vehicle_by_id.get(vid) if vid is not None else None
        if vid is not None:
            used_vehicle_ids.add(vid)
        estado = str(ov.estado) if ov is not None else "disponible"
        out.append(DotacionRowOut(
            fecha=fecha.isoformat(),
            empresa_id=int(d.empresa_id),
            empresa_nombre=empresas.get(int(d.empresa_id)),
            driver_id=str(d.driver_id),
            driver_name=str(d.name),
            driver_active=bool(d.active),
            default_vehicle_id=default_vid,
            vehicle_id=vid,
            vehicle_name=(str(vehicle.name) if vehicle is not None else d.vehicle_name),
            plate=(str(vehicle.plate) if vehicle is not None and vehicle.plate is not None else None),
            vehicle_active=bool(vehicle.active) if vehicle is not None else True,
            estado=estado,  # type: ignore[arg-type]
            motivo=str(ov.motivo) if ov is not None and ov.motivo else None,
            updated_at=(
                ov.updated_at.isoformat()
                if ov is not None and hasattr(ov.updated_at, "isoformat")
                else (str(ov.updated_at) if ov is not None and ov.updated_at is not None else None)
            ),
        ))

    for v in vehicles:
        vid = int(v.vehicle_id)
        if vid in used_vehicle_ids:
            continue
        ov = by_vehicle.get(vid)
        estado = str(ov.estado) if ov is not None else "disponible"
        out.append(DotacionRowOut(
            fecha=fecha.isoformat(),
            empresa_id=int(v.empresa_id),
            empresa_nombre=empresas.get(int(v.empresa_id)),
            driver_id=None,
            driver_name=None,
            driver_active=True,
            default_vehicle_id=vid,
            vehicle_id=vid,
            vehicle_name=str(v.name),
            plate=str(v.plate) if v.plate is not None else None,
            vehicle_active=bool(v.active),
            estado=estado,  # type: ignore[arg-type]
            motivo=str(ov.motivo) if ov is not None and ov.motivo else None,
            updated_at=(
                ov.updated_at.isoformat()
                if ov is not None and hasattr(ov.updated_at, "isoformat")
                else (str(ov.updated_at) if ov is not None and ov.updated_at is not None else None)
            ),
        ))

    out.sort(key=lambda r: (r.empresa_id, r.vehicle_id or 0, r.driver_name or ""))
    return out


def _validate_dotacion_target(cn, req: DotacionUpdate) -> None:
    if req.driver_id is None and req.vehicle_id is None:
        raise HTTPException(400, "driver_id o vehicle_id requerido")
    cur = cn.cursor()
    if req.driver_id is not None:
        cur.execute("SELECT empresa_id FROM fpoc.drivers WHERE driver_id = ?", req.driver_id)
        d = cur.fetchone()
        if not d:
            raise HTTPException(404, "driver no encontrado")
        if int(d.empresa_id) != req.empresa_id:
            raise HTTPException(400, "driver no pertenece a la empresa indicada")
    if req.vehicle_id is not None:
        cur.execute("SELECT empresa_id FROM fpoc.vehicles WHERE vehicle_id = ?", req.vehicle_id)
        v = cur.fetchone()
        if not v:
            raise HTTPException(404, "vehicle no encontrado")
        if int(v.empresa_id) != req.empresa_id:
            raise HTTPException(400, "vehiculo no pertenece a la empresa indicada")


@router.get("/dotacion-diaria", response_model=list[DotacionRowOut])
def list_dotacion_diaria(
    fecha: Optional[date] = Query(default=None),
    empresa_id: Optional[int] = Query(default=None),
    user: CurrentUser = Depends(current_user),
) -> list[DotacionRowOut]:
    fecha = fecha or date.today()
    empresa_ids = _dotacion_empresa_ids(user, empresa_id)
    for eid in empresa_ids:
        _can_access_empresa(user, eid)
    return _fetch_dotacion_rows(fecha, empresa_ids)


@router.put("/dotacion-diaria")
def upsert_dotacion_diaria(
    req: DotacionUpdate,
    user: CurrentUser = Depends(current_user),
) -> dict:
    _can_access_empresa(user, req.empresa_id)
    with get_conn() as cn:
        _validate_dotacion_target(cn, req)
        cur = cn.cursor()
        if req.driver_id is not None:
            cur.execute(
                """
                SELECT dotacion_id
                FROM fpoc.dotacion_diaria
                WHERE fecha = ? AND empresa_id = ? AND driver_id = ?
                """,
                req.fecha.isoformat(), req.empresa_id, req.driver_id,
            )
        else:
            cur.execute(
                """
                SELECT dotacion_id
                FROM fpoc.dotacion_diaria
                WHERE fecha = ? AND empresa_id = ? AND driver_id IS NULL AND vehicle_id = ?
                """,
                req.fecha.isoformat(), req.empresa_id, req.vehicle_id,
            )
        existing = cur.fetchone()
        if existing:
            cur.execute(
                """
                UPDATE fpoc.dotacion_diaria
                   SET vehicle_id = ?, estado = ?, motivo = ?,
                       updated_by_user_id = ?, updated_at = CURRENT_TIMESTAMP
                 WHERE dotacion_id = ?
                """,
                req.vehicle_id, req.estado, req.motivo, user.user_id, int(existing.dotacion_id),
            )
            dotacion_id = int(existing.dotacion_id)
        else:
            cur.execute(
                """
                INSERT INTO fpoc.dotacion_diaria
                    (fecha, empresa_id, driver_id, vehicle_id, estado, motivo, updated_by_user_id)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                req.fecha.isoformat(), req.empresa_id, req.driver_id, req.vehicle_id,
                req.estado, req.motivo, user.user_id,
            )
            dotacion_id = int(getattr(cur, "lastrowid", 0) or 0)
        cn.commit()
    return {"status": "ok", "dotacion_id": dotacion_id}


# ============================================================================
# Clients
# ============================================================================
class ClientIn(BaseModel):
    customer_id: str = Field(min_length=1, max_length=20)
    title: str = Field(min_length=1, max_length=200)
    address: str = Field(min_length=1, max_length=500)
    latitude: float = Field(ge=-90, le=90)
    longitude: float = Field(ge=-180, le=180)
    is_recurrent: bool = False
    in_problem_comuna: bool = False
    notes: Optional[str] = Field(default=None, max_length=500)


class ClientUpdate(BaseModel):
    title: Optional[str] = Field(default=None, min_length=1, max_length=200)
    address: Optional[str] = Field(default=None, min_length=1, max_length=500)
    latitude: Optional[float] = Field(default=None, ge=-90, le=90)
    longitude: Optional[float] = Field(default=None, ge=-180, le=180)
    is_recurrent: Optional[bool] = None
    in_problem_comuna: Optional[bool] = None
    notes: Optional[str] = Field(default=None, max_length=500)


class ClientOut(BaseModel):
    customer_id: str
    title: str
    address: str
    latitude: float
    longitude: float
    is_recurrent: bool
    in_problem_comuna: bool
    notes: Optional[str] = None


class ClientsPage(BaseModel):
    rows: list[ClientOut]
    total: int
    limit: int
    offset: int


def _client_row(r) -> ClientOut:
    return ClientOut(
        customer_id=r.customer_id, title=r.title, address=r.address,
        latitude=float(r.latitude), longitude=float(r.longitude),
        is_recurrent=bool(r.is_recurrent),
        in_problem_comuna=bool(r.in_problem_comuna),
        notes=r.notes,
    )


@router.get("/clients", response_model=ClientsPage)
def list_clients(
    limit: int = Query(default=100, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    search: Optional[str] = Query(default=None),
    only_recurrent: bool = Query(default=False),
    only_problem: bool = Query(default=False),
    _: CurrentUser = Depends(require_admin),
) -> ClientsPage:
    where, params = ["1=1"], []
    if search:
        where.append("(title LIKE ? OR customer_id LIKE ? OR address LIKE ?)")
        like = f"%{search}%"; params.extend([like, like, like])
    if only_recurrent:
        where.append("is_recurrent = 1")
    if only_problem:
        where.append("in_problem_comuna = 1")
    where_sql = " AND ".join(where)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(f"SELECT COUNT(*) FROM fpoc.clients WHERE {where_sql}", *params)
        total = int(cur.fetchone()[0])
        cur.execute(
            f"""SELECT customer_id, title, address, latitude, longitude,
                       is_recurrent, in_problem_comuna, notes
                FROM fpoc.clients WHERE {where_sql}
                ORDER BY title
                LIMIT ? OFFSET ?""",
            *params, limit, offset,
        )
        rows = [_client_row(r) for r in cur.fetchall()]
    return ClientsPage(rows=rows, total=total, limit=limit, offset=offset)


@router.post("/clients", response_model=ClientOut)
def create_client(req: ClientIn, _: CurrentUser = Depends(require_admin)) -> ClientOut:
    with get_conn() as cn:
        cur = cn.cursor()
        try:
            cur.execute(
                """INSERT INTO fpoc.clients
                    (customer_id, title, address, latitude, longitude,
                     is_recurrent, in_problem_comuna, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                req.customer_id, req.title, req.address, req.latitude, req.longitude,
                1 if req.is_recurrent else 0,
                1 if req.in_problem_comuna else 0,
                req.notes,
            )
            cn.commit()
        except Exception as e:  # noqa: BLE001
            cn.rollback()
            raise HTTPException(409, f"customer_id duplicado: {e}")
        cur.execute(
            """SELECT customer_id, title, address, latitude, longitude,
                      is_recurrent, in_problem_comuna, notes
               FROM fpoc.clients WHERE customer_id = ?""",
            req.customer_id,
        )
        out = _client_row(cur.fetchone())
    _refresh_state_maestros()
    return out


@router.put("/clients/{customer_id}", response_model=ClientOut)
def update_client(customer_id: str, req: ClientUpdate,
                   _: CurrentUser = Depends(require_admin)) -> ClientOut:
    sets, params = [], []
    for field in ["title", "address", "latitude", "longitude", "notes"]:
        v = getattr(req, field)
        if v is not None:
            sets.append(f"{field} = ?"); params.append(v)
    if req.is_recurrent is not None:
        sets.append("is_recurrent = ?"); params.append(1 if req.is_recurrent else 0)
    if req.in_problem_comuna is not None:
        sets.append("in_problem_comuna = ?"); params.append(1 if req.in_problem_comuna else 0)
    sets.append("updated_at = CURRENT_TIMESTAMP")
    if not sets:
        raise HTTPException(400, "nada que actualizar")
    params.append(customer_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(f"UPDATE fpoc.clients SET {', '.join(sets)} WHERE customer_id = ?", *params)
        if cur.rowcount == 0:
            raise HTTPException(404, "client no encontrado")
        cn.commit()
        cur.execute(
            """SELECT customer_id, title, address, latitude, longitude,
                      is_recurrent, in_problem_comuna, notes
               FROM fpoc.clients WHERE customer_id = ?""",
            customer_id,
        )
        out = _client_row(cur.fetchone())
    _refresh_state_maestros()
    return out


@router.delete("/clients/{customer_id}")
def delete_client(customer_id: str, _: CurrentUser = Depends(require_admin)) -> dict:
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute("DELETE FROM fpoc.clients WHERE customer_id = ?", customer_id)
        if cur.rowcount == 0:
            raise HTTPException(404, "client no encontrado")
        cn.commit()
    _refresh_state_maestros()
    return {"deleted": customer_id}


# ============================================================================
# Excel template + upload masivo (drivers, vehicles, dotacion)
# ============================================================================
class BulkUploadResult(BaseModel):
    created: int = 0
    updated: int = 0
    errors: list[str] = []


def _xlsx_response(filename: str, headers: list[str], rows: list[list[Any]]) -> StreamingResponse:
    """Genera un xlsx in-memory con headers + rows. Usa openpyxl (ya en deps)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    # Formatear header en negrita
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _read_xlsx_rows(file_bytes: bytes) -> tuple[list[str], list[list[Any]]]:
    """Devuelve (headers, rows). Headers normalizados a lower."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip().lower() if h is not None else "" for h in headers_raw]
    rows = [list(r) for r in rows_iter if any(c is not None for c in r)]
    return headers, rows


# ----- Drivers -----
@router.get("/drivers/template")
def drivers_template(empresa_id: int = Query(...),
                      user: CurrentUser = Depends(require_fleet_access)) -> StreamingResponse:
    _enforce_fleet_empresa(user, empresa_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            """SELECT driver_id, name, phone, license, vehicle_id, vehicle_name, active
                FROM fpoc.drivers WHERE empresa_id = ? ORDER BY driver_id""",
            empresa_id,
        )
        rows = [
            [r.driver_id, r.name, r.phone or "", r.license or "",
             int(r.vehicle_id), r.vehicle_name or "", 1 if r.active else 0]
            for r in cur.fetchall()
        ]
    headers = ["driver_id", "name", "phone", "license", "vehicle_id", "vehicle_name", "active"]
    return _xlsx_response(f"drivers_empresa_{empresa_id}.xlsx", headers, rows)


@router.post("/drivers/upload", response_model=BulkUploadResult)
async def drivers_upload(
    empresa_id: int = Query(...),
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_fleet_access),
) -> BulkUploadResult:
    _enforce_fleet_empresa(user, empresa_id)
    headers, rows = _read_xlsx_rows(await file.read())
    required = ["driver_id", "name", "vehicle_id", "vehicle_name"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(400, f"Faltan columnas requeridas: {missing}")
    idx = {h: i for i, h in enumerate(headers)}
    result = BulkUploadResult()
    with get_conn() as cn:
        cur = cn.cursor()
        for line_num, r in enumerate(rows, start=2):  # row 1 es header
            try:
                driver_id = str(r[idx["driver_id"]]).strip()
                if not driver_id:
                    continue
                name = str(r[idx["name"]] or "").strip()
                phone = str(r[idx["phone"]] or "").strip() or None if "phone" in idx else None
                lic = str(r[idx["license"]] or "").strip() or None if "license" in idx else None
                vehicle_id = int(r[idx["vehicle_id"]])
                vehicle_name = str(r[idx["vehicle_name"]] or "").strip()
                active = bool(int(r[idx["active"]] or 1)) if "active" in idx else True
                # Existe? -> update, sino insert
                cur.execute("SELECT 1 FROM fpoc.drivers WHERE driver_id = ?", driver_id)
                if cur.fetchone():
                    cur.execute(
                        """UPDATE fpoc.drivers SET name=?, phone=?, license=?, empresa_id=?,
                                  vehicle_id=?, vehicle_name=?, active=?, updated_at=CURRENT_TIMESTAMP
                                WHERE driver_id=?""",
                        name, phone, lic, empresa_id, vehicle_id, vehicle_name,
                        1 if active else 0, driver_id,
                    )
                    result.updated += 1
                else:
                    cur.execute(
                        """INSERT INTO fpoc.drivers
                            (driver_id, name, phone, license, empresa_id, vehicle_id, vehicle_name, active)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                        driver_id, name, phone, lic, empresa_id, vehicle_id, vehicle_name,
                        1 if active else 0,
                    )
                    result.created += 1
            except Exception as e:  # noqa: BLE001
                result.errors.append(f"fila {line_num}: {type(e).__name__}: {e}")
        cn.commit()
    _refresh_state_maestros()
    return result


# ----- Vehicles -----
@router.get("/vehicles/template")
def vehicles_template(empresa_id: int = Query(...),
                       user: CurrentUser = Depends(require_fleet_access)) -> StreamingResponse:
    _enforce_fleet_empresa(user, empresa_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            """SELECT vehicle_id, name, type, plate, capacity_m3, driver_id, driver_name,
                       depot_lat, depot_lon, year, active
                FROM fpoc.vehicles WHERE empresa_id = ? ORDER BY vehicle_id""",
            empresa_id,
        )
        rows = [
            [int(r.vehicle_id), r.name, r.type or "", r.plate or "",
             int(r.capacity_m3 or 0), r.driver_id or "", r.driver_name or "",
             float(r.depot_lat), float(r.depot_lon),
             int(r.year) if r.year is not None else "",
             1 if r.active else 0]
            for r in cur.fetchall()
        ]
    headers = ["vehicle_id", "name", "type", "plate", "capacity_m3",
               "driver_id", "driver_name", "depot_lat", "depot_lon", "year", "active"]
    return _xlsx_response(f"vehicles_empresa_{empresa_id}.xlsx", headers, rows)


@router.post("/vehicles/upload", response_model=BulkUploadResult)
async def vehicles_upload(
    empresa_id: int = Query(...),
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_fleet_access),
) -> BulkUploadResult:
    _enforce_fleet_empresa(user, empresa_id)
    headers, rows = _read_xlsx_rows(await file.read())
    required = ["vehicle_id", "name", "type", "plate", "capacity_m3"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(400, f"Faltan columnas requeridas: {missing}")
    idx = {h: i for i, h in enumerate(headers)}
    result = BulkUploadResult()
    with get_conn() as cn:
        cur = cn.cursor()
        for line_num, r in enumerate(rows, start=2):
            try:
                vid = int(r[idx["vehicle_id"]])
                name = str(r[idx["name"]] or "").strip()
                vtype = str(r[idx["type"]] or "").strip()
                plate = str(r[idx["plate"]] or "").strip()
                cap = int(r[idx["capacity_m3"]] or 0)
                drv_id = str(r[idx["driver_id"]] or "").strip() or None if "driver_id" in idx else None
                drv_name = str(r[idx["driver_name"]] or "").strip() or None if "driver_name" in idx else None
                lat = float(r[idx["depot_lat"]] or -33.45) if "depot_lat" in idx else -33.45
                lon = float(r[idx["depot_lon"]] or -70.66) if "depot_lon" in idx else -70.66
                year = int(r[idx["year"]]) if "year" in idx and r[idx["year"]] not in (None, "") else None
                active = bool(int(r[idx["active"]] or 1)) if "active" in idx else True
                cur.execute("SELECT 1 FROM fpoc.vehicles WHERE vehicle_id = ?", vid)
                if cur.fetchone():
                    cur.execute(
                        """UPDATE fpoc.vehicles SET empresa_id=?, name=?, type=?, plate=?,
                                  capacity_m3=?, driver_id=?, driver_name=?, depot_lat=?,
                                  depot_lon=?, year=?, active=?, updated_at=CURRENT_TIMESTAMP
                                WHERE vehicle_id=?""",
                        empresa_id, name, vtype, plate, cap, drv_id, drv_name, lat, lon,
                        year, 1 if active else 0, vid,
                    )
                    result.updated += 1
                else:
                    cur.execute(
                        """INSERT INTO fpoc.vehicles
                            (vehicle_id, empresa_id, name, type, plate, capacity_m3, driver_id,
                             driver_name, depot_lat, depot_lon, year, active)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        vid, empresa_id, name, vtype, plate, cap, drv_id, drv_name, lat, lon,
                        year, 1 if active else 0,
                    )
                    result.created += 1
            except Exception as e:  # noqa: BLE001
                result.errors.append(f"fila {line_num}: {type(e).__name__}: {e}")
        cn.commit()
    _refresh_state_maestros()
    return result


# ----- Dotación -----
@router.get("/dotacion-diaria/template")
def dotacion_template(
    fecha: date = Query(...),
    empresa_id: Optional[int] = Query(default=None),
    user: CurrentUser = Depends(current_user),
) -> StreamingResponse:
    """Template pre-rellenado con drivers de la(s) empresa(s) y su estado actual."""
    empresa_ids = _dotacion_empresa_ids(user, empresa_id)
    for eid in empresa_ids:
        _can_access_empresa(user, eid)
    rows_data = _fetch_dotacion_rows(fecha, empresa_ids)
    headers = ["empresa_id", "driver_id", "driver_name", "vehicle_id", "estado", "motivo"]
    rows = [
        [r.empresa_id, r.driver_id or "", r.driver_name or "",
         r.vehicle_id if r.vehicle_id is not None else "",
         r.estado, r.motivo or ""]
        for r in rows_data
    ]
    suffix = f"empresa_{empresa_id}" if empresa_id else "todas"
    return _xlsx_response(f"dotacion_{fecha.isoformat()}_{suffix}.xlsx", headers, rows)


@router.post("/dotacion-diaria/upload", response_model=BulkUploadResult)
async def dotacion_upload(
    fecha: date = Query(...),
    file: UploadFile = File(...),
    user: CurrentUser = Depends(current_user),
) -> BulkUploadResult:
    headers, rows = _read_xlsx_rows(await file.read())
    required = ["empresa_id", "driver_id", "estado"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(400, f"Faltan columnas requeridas: {missing}")
    idx = {h: i for i, h in enumerate(headers)}
    valid_estados = {"disponible", "ausente", "licencia", "mantencion", "baja", "reemplazo"}
    result = BulkUploadResult()
    with get_conn() as cn:
        cur = cn.cursor()
        for line_num, r in enumerate(rows, start=2):
            try:
                eid = int(r[idx["empresa_id"]])
                _can_access_empresa(user, eid)
                drv_id = str(r[idx["driver_id"]]).strip() or None
                vid = (int(r[idx["vehicle_id"]])
                        if "vehicle_id" in idx and r[idx["vehicle_id"]] not in (None, "") else None)
                estado = str(r[idx["estado"]] or "disponible").strip().lower()
                if estado not in valid_estados:
                    raise ValueError(f"estado inválido: {estado}")
                motivo = str(r[idx["motivo"]] or "").strip() or None if "motivo" in idx else None
                if drv_id is None and vid is None:
                    raise ValueError("driver_id o vehicle_id requerido")
                # Buscar override existente
                if drv_id:
                    cur.execute(
                        "SELECT dotacion_id FROM fpoc.dotacion_diaria WHERE fecha=? AND empresa_id=? AND driver_id=?",
                        fecha.isoformat(), eid, drv_id,
                    )
                else:
                    cur.execute(
                        "SELECT dotacion_id FROM fpoc.dotacion_diaria WHERE fecha=? AND empresa_id=? AND vehicle_id=?",
                        fecha.isoformat(), eid, vid,
                    )
                existing = cur.fetchone()
                if existing:
                    cur.execute(
                        """UPDATE fpoc.dotacion_diaria
                                SET vehicle_id=?, estado=?, motivo=?, updated_by_user_id=?, updated_at=CURRENT_TIMESTAMP
                              WHERE dotacion_id=?""",
                        vid, estado, motivo, user.user_id, int(existing.dotacion_id),
                    )
                    result.updated += 1
                else:
                    cur.execute(
                        """INSERT INTO fpoc.dotacion_diaria
                            (fecha, empresa_id, driver_id, vehicle_id, estado, motivo, updated_by_user_id)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        fecha.isoformat(), eid, drv_id, vid, estado, motivo, user.user_id,
                    )
                    result.created += 1
            except HTTPException:
                raise
            except Exception as e:  # noqa: BLE001
                result.errors.append(f"fila {line_num}: {type(e).__name__}: {e}")
        cn.commit()
    return result


# ============================================================================
# Driver documents
# ============================================================================
DOC_TIPOS = ("licencia", "antecedentes", "contrato", "poliza", "certificacion", "otro")


class DriverDocOut(BaseModel):
    doc_id: int
    driver_id: str
    tipo: str
    filename: str
    file_size: int
    content_type: Optional[str] = None
    uploaded_at: str
    uploaded_by_user_id: Optional[int] = None
    expires_at: Optional[str] = None
    notes: Optional[str] = None


def _doc_row(r) -> DriverDocOut:
    return DriverDocOut(
        doc_id=int(r.doc_id),
        driver_id=str(r.driver_id),
        tipo=str(r.tipo),
        filename=str(r.filename),
        file_size=int(r.file_size or 0),
        content_type=r.content_type,
        uploaded_at=(r.uploaded_at.isoformat() if hasattr(r.uploaded_at, "isoformat") else str(r.uploaded_at)),
        uploaded_by_user_id=int(r.uploaded_by_user_id) if r.uploaded_by_user_id is not None else None,
        expires_at=(r.expires_at.isoformat() if hasattr(r.expires_at, "isoformat") else (str(r.expires_at) if r.expires_at else None)),
        notes=r.notes,
    )


def _enforce_driver_access(user: CurrentUser, driver_id: str) -> None:
    """transport_manager solo puede ver/tocar drivers de su empresa."""
    if user.is_falabella:
        return
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute("SELECT empresa_id FROM fpoc.drivers WHERE driver_id = ?", driver_id)
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "driver no encontrado")
        _enforce_fleet_empresa(user, int(row.empresa_id) if row.empresa_id is not None else None)


@router.get("/drivers/{driver_id}/documents", response_model=list[DriverDocOut])
def list_driver_documents(
    driver_id: str,
    user: CurrentUser = Depends(require_fleet_access),
) -> list[DriverDocOut]:
    _enforce_driver_access(user, driver_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            """SELECT doc_id, driver_id, tipo, filename, file_size, content_type,
                       uploaded_at, uploaded_by_user_id, expires_at, notes
                FROM fpoc.driver_documents
                WHERE driver_id = ?
                ORDER BY uploaded_at DESC""",
            driver_id,
        )
        return [_doc_row(r) for r in cur.fetchall()]


@router.post("/drivers/{driver_id}/documents", response_model=DriverDocOut)
async def upload_driver_document(
    driver_id: str,
    tipo: str = Query(..., description=f"Uno de: {', '.join(DOC_TIPOS)}"),
    expires_at: Optional[str] = Query(default=None, description="YYYY-MM-DD opcional"),
    notes: Optional[str] = Query(default=None, max_length=500),
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_fleet_access),
) -> DriverDocOut:
    _enforce_driver_access(user, driver_id)
    if tipo not in DOC_TIPOS:
        raise HTTPException(400, f"tipo inválido. Permitidos: {DOC_TIPOS}")
    data = await file.read()
    if not data:
        raise HTTPException(400, "archivo vacío")
    if len(data) > 25 * 1024 * 1024:
        raise HTTPException(413, "archivo > 25MB")

    import uuid as _uuid
    from storage import upload as storage_upload
    safe_name = (file.filename or "documento").replace("/", "_").replace("\\", "_")
    blob_path = f"drivers/{driver_id}/{_uuid.uuid4().hex}_{safe_name}"
    try:
        storage_upload(blob_path, data, content_type=file.content_type)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(500, f"storage upload falló: {e}")

    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            """INSERT INTO fpoc.driver_documents
                (driver_id, tipo, filename, blob_path, file_size, content_type,
                 uploaded_by_user_id, expires_at, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            driver_id, tipo, safe_name, blob_path, len(data),
            file.content_type, user.user_id, expires_at, notes,
        )
        cn.commit()
        cur.execute(
            """SELECT doc_id, driver_id, tipo, filename, file_size, content_type,
                       uploaded_at, uploaded_by_user_id, expires_at, notes
                FROM fpoc.driver_documents
                WHERE driver_id = ? AND blob_path = ?""",
            driver_id, blob_path,
        )
        return _doc_row(cur.fetchone())


@router.get("/drivers/{driver_id}/documents/{doc_id}/download")
def download_driver_document(
    driver_id: str,
    doc_id: int,
    user: CurrentUser = Depends(require_fleet_access),
) -> StreamingResponse:
    _enforce_driver_access(user, driver_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            "SELECT filename, blob_path, content_type FROM fpoc.driver_documents "
            "WHERE driver_id = ? AND doc_id = ?",
            driver_id, doc_id,
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "documento no encontrado")
    from storage import download as storage_download
    try:
        data, ct = storage_download(str(row.blob_path))
    except FileNotFoundError:
        raise HTTPException(404, "blob ausente en storage")
    content_type = row.content_type or ct or "application/octet-stream"
    return StreamingResponse(
        io.BytesIO(data),
        media_type=content_type,
        headers={"Content-Disposition": f'attachment; filename="{row.filename}"'},
    )


@router.delete("/drivers/{driver_id}/documents/{doc_id}")
def delete_driver_document(
    driver_id: str,
    doc_id: int,
    user: CurrentUser = Depends(require_fleet_access),
) -> dict:
    _enforce_driver_access(user, driver_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            "SELECT blob_path FROM fpoc.driver_documents WHERE driver_id = ? AND doc_id = ?",
            driver_id, doc_id,
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "documento no encontrado")
        blob_path = str(row.blob_path)
        cur.execute("DELETE FROM fpoc.driver_documents WHERE doc_id = ?", doc_id)
        cn.commit()
    from storage import delete as storage_delete
    storage_delete(blob_path)
    return {"deleted": doc_id}


# ============================================================================
# Capacitaciones — catálogo de módulos + registros por driver
# ============================================================================
class CapacitacionModuloIn(BaseModel):
    codigo: str = Field(min_length=1, max_length=50)
    nombre: str = Field(min_length=1, max_length=200)
    descripcion: Optional[str] = Field(default=None, max_length=500)
    validez_meses: int = Field(ge=1, le=120, default=12)
    activo: bool = True


class CapacitacionModuloUpdate(BaseModel):
    nombre: Optional[str] = Field(default=None, min_length=1, max_length=200)
    descripcion: Optional[str] = Field(default=None, max_length=500)
    validez_meses: Optional[int] = Field(default=None, ge=1, le=120)
    activo: Optional[bool] = None


class CapacitacionModuloOut(BaseModel):
    modulo_id: int
    codigo: str
    nombre: str
    descripcion: Optional[str] = None
    validez_meses: int
    activo: bool


def _modulo_row(r) -> CapacitacionModuloOut:
    return CapacitacionModuloOut(
        modulo_id=int(r.modulo_id),
        codigo=str(r.codigo),
        nombre=str(r.nombre),
        descripcion=r.descripcion,
        validez_meses=int(r.validez_meses),
        activo=bool(r.activo),
    )


@router.get("/capacitacion-modulos", response_model=list[CapacitacionModuloOut])
def list_capacitacion_modulos(
    only_active: bool = Query(default=False),
    user: CurrentUser = Depends(current_user),
) -> list[CapacitacionModuloOut]:
    where = "WHERE activo = 1" if only_active else ""
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            f"""SELECT modulo_id, codigo, nombre, descripcion, validez_meses, activo
                FROM fpoc.capacitacion_modulos
                {where}
                ORDER BY nombre"""
        )
        return [_modulo_row(r) for r in cur.fetchall()]


@router.post("/capacitacion-modulos", response_model=CapacitacionModuloOut)
def create_capacitacion_modulo(
    req: CapacitacionModuloIn,
    _: CurrentUser = Depends(require_admin),
) -> CapacitacionModuloOut:
    with get_conn() as cn:
        cur = cn.cursor()
        try:
            cur.execute(
                """INSERT INTO fpoc.capacitacion_modulos
                    (codigo, nombre, descripcion, validez_meses, activo)
                   VALUES (?, ?, ?, ?, ?)""",
                req.codigo, req.nombre, req.descripcion, req.validez_meses,
                1 if req.activo else 0,
            )
            cn.commit()
        except Exception as e:  # noqa: BLE001
            cn.rollback()
            raise HTTPException(409, f"código duplicado o datos inválidos: {e}")
        cur.execute(
            """SELECT modulo_id, codigo, nombre, descripcion, validez_meses, activo
                FROM fpoc.capacitacion_modulos WHERE codigo = ?""",
            req.codigo,
        )
        return _modulo_row(cur.fetchone())


@router.put("/capacitacion-modulos/{modulo_id}", response_model=CapacitacionModuloOut)
def update_capacitacion_modulo(
    modulo_id: int,
    req: CapacitacionModuloUpdate,
    _: CurrentUser = Depends(require_admin),
) -> CapacitacionModuloOut:
    sets, params = [], []
    if req.nombre is not None:
        sets.append("nombre = ?"); params.append(req.nombre)
    if req.descripcion is not None:
        sets.append("descripcion = ?"); params.append(req.descripcion)
    if req.validez_meses is not None:
        sets.append("validez_meses = ?"); params.append(req.validez_meses)
    if req.activo is not None:
        sets.append("activo = ?"); params.append(1 if req.activo else 0)
    if not sets:
        raise HTTPException(400, "nada que actualizar")
    params.append(modulo_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(f"UPDATE fpoc.capacitacion_modulos SET {', '.join(sets)} WHERE modulo_id = ?", *params)
        if cur.rowcount == 0:
            raise HTTPException(404, "módulo no encontrado")
        cn.commit()
        cur.execute(
            """SELECT modulo_id, codigo, nombre, descripcion, validez_meses, activo
                FROM fpoc.capacitacion_modulos WHERE modulo_id = ?""",
            modulo_id,
        )
        return _modulo_row(cur.fetchone())


@router.delete("/capacitacion-modulos/{modulo_id}")
def delete_capacitacion_modulo(
    modulo_id: int,
    _: CurrentUser = Depends(require_admin),
) -> dict:
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute("DELETE FROM fpoc.capacitacion_modulos WHERE modulo_id = ?", modulo_id)
        if cur.rowcount == 0:
            raise HTTPException(404, "módulo no encontrado")
        cn.commit()
    return {"deleted": modulo_id}


# ----- Registros por driver -----
class DriverCapacitacionIn(BaseModel):
    modulo_id: int = Field(ge=1)
    fecha_completado: date
    vence_at: Optional[date] = None  # si null, calculamos = fecha + validez_meses
    notas: Optional[str] = Field(default=None, max_length=500)
    doc_id: Optional[int] = None


class DriverCapacitacionUpdate(BaseModel):
    fecha_completado: Optional[date] = None
    vence_at: Optional[date] = None
    notas: Optional[str] = Field(default=None, max_length=500)
    doc_id: Optional[int] = None


class DriverCapacitacionOut(BaseModel):
    cap_id: int
    driver_id: str
    modulo_id: int
    modulo_codigo: str
    modulo_nombre: str
    fecha_completado: str
    vence_at: Optional[str] = None
    notas: Optional[str] = None
    doc_id: Optional[int] = None
    created_by: Optional[int] = None
    created_at: str


def _cap_row(r) -> DriverCapacitacionOut:
    def _iso(v):
        if v is None:
            return None
        return v.isoformat() if hasattr(v, "isoformat") else str(v)
    return DriverCapacitacionOut(
        cap_id=int(r.cap_id),
        driver_id=str(r.driver_id),
        modulo_id=int(r.modulo_id),
        modulo_codigo=str(r.modulo_codigo),
        modulo_nombre=str(r.modulo_nombre),
        fecha_completado=_iso(r.fecha_completado) or "",
        vence_at=_iso(r.vence_at),
        notas=r.notas,
        doc_id=int(r.doc_id) if r.doc_id is not None else None,
        created_by=int(r.created_by) if r.created_by is not None else None,
        created_at=_iso(r.created_at) or "",
    )


def _compute_vence_at(fecha_completado: date, validez_meses: int) -> date:
    """Suma N meses a la fecha (manejando fin de mes)."""
    from calendar import monthrange
    y = fecha_completado.year
    m = fecha_completado.month + validez_meses
    while m > 12:
        m -= 12
        y += 1
    last_day = monthrange(y, m)[1]
    d = min(fecha_completado.day, last_day)
    return date(y, m, d)


@router.get("/drivers/{driver_id}/capacitaciones", response_model=list[DriverCapacitacionOut])
def list_driver_capacitaciones(
    driver_id: str,
    user: CurrentUser = Depends(require_fleet_access),
) -> list[DriverCapacitacionOut]:
    _enforce_driver_access(user, driver_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            """SELECT c.cap_id, c.driver_id, c.modulo_id,
                       m.codigo AS modulo_codigo, m.nombre AS modulo_nombre,
                       c.fecha_completado, c.vence_at, c.notas, c.doc_id,
                       c.created_by, c.created_at
                FROM fpoc.driver_capacitaciones c
                INNER JOIN fpoc.capacitacion_modulos m ON m.modulo_id = c.modulo_id
                WHERE c.driver_id = ?
                ORDER BY c.fecha_completado DESC""",
            driver_id,
        )
        return [_cap_row(r) for r in cur.fetchall()]


@router.post("/drivers/{driver_id}/capacitaciones", response_model=DriverCapacitacionOut)
def create_driver_capacitacion(
    driver_id: str,
    req: DriverCapacitacionIn,
    user: CurrentUser = Depends(require_fleet_access),
) -> DriverCapacitacionOut:
    _enforce_driver_access(user, driver_id)
    # Calcular vence_at si no vino
    vence_at = req.vence_at
    if vence_at is None:
        with get_conn() as cn:
            cur = cn.cursor()
            cur.execute(
                "SELECT validez_meses FROM fpoc.capacitacion_modulos WHERE modulo_id = ?",
                req.modulo_id,
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(400, "módulo inválido")
            vence_at = _compute_vence_at(req.fecha_completado, int(row.validez_meses))

    with get_conn() as cn:
        cur = cn.cursor()
        try:
            cur.execute(
                """INSERT INTO fpoc.driver_capacitaciones
                    (driver_id, modulo_id, fecha_completado, vence_at, notas, doc_id, created_by)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                driver_id, req.modulo_id,
                req.fecha_completado.isoformat(), vence_at.isoformat() if vence_at else None,
                req.notas, req.doc_id, user.user_id,
            )
            cn.commit()
        except Exception as e:  # noqa: BLE001
            cn.rollback()
            raise HTTPException(400, f"error guardando capacitación: {e}")
        cur.execute(
            """SELECT TOP 1 c.cap_id, c.driver_id, c.modulo_id,
                       m.codigo AS modulo_codigo, m.nombre AS modulo_nombre,
                       c.fecha_completado, c.vence_at, c.notas, c.doc_id,
                       c.created_by, c.created_at
                FROM fpoc.driver_capacitaciones c
                INNER JOIN fpoc.capacitacion_modulos m ON m.modulo_id = c.modulo_id
                WHERE c.driver_id = ? AND c.modulo_id = ?
                ORDER BY c.cap_id DESC""",
            driver_id, req.modulo_id,
        )
        return _cap_row(cur.fetchone())


@router.put("/drivers/{driver_id}/capacitaciones/{cap_id}", response_model=DriverCapacitacionOut)
def update_driver_capacitacion(
    driver_id: str,
    cap_id: int,
    req: DriverCapacitacionUpdate,
    user: CurrentUser = Depends(require_fleet_access),
) -> DriverCapacitacionOut:
    _enforce_driver_access(user, driver_id)
    sets, params = [], []
    if req.fecha_completado is not None:
        sets.append("fecha_completado = ?"); params.append(req.fecha_completado.isoformat())
    if req.vence_at is not None:
        sets.append("vence_at = ?"); params.append(req.vence_at.isoformat())
    if req.notas is not None:
        sets.append("notas = ?"); params.append(req.notas)
    if req.doc_id is not None:
        sets.append("doc_id = ?"); params.append(req.doc_id)
    if not sets:
        raise HTTPException(400, "nada que actualizar")
    params += [cap_id, driver_id]
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            f"UPDATE fpoc.driver_capacitaciones SET {', '.join(sets)} "
            f"WHERE cap_id = ? AND driver_id = ?",
            *params,
        )
        if cur.rowcount == 0:
            raise HTTPException(404, "capacitación no encontrada")
        cn.commit()
        cur.execute(
            """SELECT c.cap_id, c.driver_id, c.modulo_id,
                       m.codigo AS modulo_codigo, m.nombre AS modulo_nombre,
                       c.fecha_completado, c.vence_at, c.notas, c.doc_id,
                       c.created_by, c.created_at
                FROM fpoc.driver_capacitaciones c
                INNER JOIN fpoc.capacitacion_modulos m ON m.modulo_id = c.modulo_id
                WHERE c.cap_id = ?""",
            cap_id,
        )
        return _cap_row(cur.fetchone())


@router.delete("/drivers/{driver_id}/capacitaciones/{cap_id}")
def delete_driver_capacitacion(
    driver_id: str,
    cap_id: int,
    user: CurrentUser = Depends(require_fleet_access),
) -> dict:
    _enforce_driver_access(user, driver_id)
    with get_conn() as cn:
        cur = cn.cursor()
        cur.execute(
            "DELETE FROM fpoc.driver_capacitaciones WHERE cap_id = ? AND driver_id = ?",
            cap_id, driver_id,
        )
        if cur.rowcount == 0:
            raise HTTPException(404, "capacitación no encontrada")
        cn.commit()
    return {"deleted": cap_id}
