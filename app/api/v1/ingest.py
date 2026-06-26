"""Falabella XLSX/JSON ingest pipeline — CR-019 (bulk fix CR-019.1).

Endpoints:
  POST /api/v1/ingest/falabella-xlsx   multipart .xlsx upload
  POST /api/v1/ingest/falabella-json   JSON body {rows: [FalabellaRow, ...]}

Performance design (CR-019.1):
  The first version did `session.add(obj); await session.flush()` PER row and
  one SELECT-per-key for idempotency lookups. Against Azure SQL (50-100ms RTT)
  that meant ~6 min just on roundtrips for the 2.2k-row reference file, well
  past Azure's request timeout.

  This rewrite:
   1. Reads ALL rows from XLSX/JSON, validates with Pydantic.
   2. Pre-fetches existing keys per empresa in ONE query each:
        - clientes by surrogate `FAL-{do}`
        - rutas by `folio`
        - visitas by (ruta_id, folio_cliente)
   3. Bulk-inserts via `session.execute(insert(Model), [dict, ...])` in
      chunks of 500 with `commit()` per chunk so the engine can use pyodbc
      `fast_executemany` (enabled in `app/db/session.py`).
   4. After clientes are inserted, re-SELECTs them to recover their generated
      PKs and build a `{(empresa_id, surrogate_rut): cliente_id}` map (rutas
      idem with `folio`).
   5. Visitas inserts run last and reference the cached PKs — no per-row
      lookups during the visita pass.
   6. Geocoding runs as a fire-and-forget `asyncio.create_task()` AFTER the
      final commit; opens its own session.
"""
from __future__ import annotations

import asyncio
import io
import secrets
import time
from collections import defaultdict
from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import func, insert, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.geocoding import centroide_comuna, geocode_pending_clientes
from app.core.security import current_user
from app.db.models.cliente import Cliente
from app.db.models.dia_operativo import DiaOperativo
from app.db.models.driver import Driver
from app.db.models.empresa import Empresa
from app.db.models.ruta import Ruta
from app.db.models.user import User
from app.db.models.vehicle import Vehicle
from app.db.models.visita import Visita
from app.db.session import get_db
from app.schemas.ingest import FalabellaJSONBody, FalabellaRow, IngestResult

router = APIRouter(prefix="/api/v1/ingest", tags=["ingest"])


# ── Mappings ────────────────────────────────────────────────────────────────

EMPRESA_MAP: dict[int, int] = {22: 1, 23: 2, 25: 3, 27: 4, 33: 5}

ESTADO_MAP: dict[str, str] = {
    "terminado": "entregado",
    "entregado": "entregado",
    "no entregado": "no_entregado",
    "no_entregado": "no_entregado",
}

# Chunk size for bulk inserts. Tuned for Azure SQL request-packet limits
# (default ~4 KB packet, ~10 MB statement) and pyodbc fast_executemany.
BULK_CHUNK = 500


# ── Helpers ─────────────────────────────────────────────────────────────────


def _require_falabella_role(user: User) -> None:
    if user.role not in ("falabella_admin", "falabella_ops"):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Only Falabella admin/ops can ingest source data",
        )


def _parse_fechainicioruta(raw: Any) -> datetime:
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date):
        return datetime.combine(raw, datetime.min.time())
    s = str(raw).strip()
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError as e:
            raise ValueError(f"Cannot parse fechainicioruta='{raw}'") from e


def _parse_fechapactada(raw: Any) -> date | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, date) and not isinstance(raw, datetime):
        return raw
    if isinstance(raw, datetime):
        return raw.date()
    s = str(raw).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").date()
        except ValueError:
            return None


def _map_estado(raw: str | None) -> str:
    if not raw:
        return "pendiente"
    return ESTADO_MAP.get(str(raw).strip().lower(), "pendiente")


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Read XLSX, prefer sheet 'Geo'. Row 1 = headers, rest = data."""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, f"Cannot read XLSX: {e}"
        ) from None

    ws = wb["Geo"] if "Geo" in wb.sheetnames else wb.active
    if ws is None:
        wb.close()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "XLSX has no readable sheets")

    headers: list[str] = []
    rows: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h is not None else f"col_{j}" for j, h in enumerate(row)]
            continue
        if all(c is None for c in row):
            continue
        rows.append(dict(zip(headers, row, strict=False)))
    wb.close()
    return rows


async def _chunked_insert(
    db: AsyncSession,
    model: type,
    payload: list[dict[str, Any]],
    label: str,
) -> int:
    """Bulk insert `payload` into `model` in chunks of BULK_CHUNK, commit per chunk."""
    if not payload:
        return 0
    total = len(payload)
    t0 = time.perf_counter()
    inserted = 0
    for i in range(0, total, BULK_CHUNK):
        chunk = payload[i : i + BULK_CHUNK]
        await db.execute(insert(model), chunk)
        await db.commit()
        inserted += len(chunk)
        logger.info(
            f"  {label}: inserted {inserted}/{total} "
            f"(chunk {i // BULK_CHUNK + 1}, {time.perf_counter() - t0:.1f}s)"
        )
    return inserted


# ── Get-or-create for low-volume entities (drivers, vehicles, dias) ─────────


async def _ensure_dia(
    db: AsyncSession, empresa_id: int, fecha: date, user: User
) -> DiaOperativo:
    existing = (
        await db.execute(
            select(DiaOperativo).where(
                DiaOperativo.empresa_id == empresa_id,
                DiaOperativo.fecha == fecha,
            )
        )
    ).scalar_one_or_none()
    if existing is not None:
        return existing
    dia = DiaOperativo(
        empresa_id=empresa_id,
        fecha=fecha,
        estado="BORRADOR",
        created_by_user_id=user.user_id,
        notas="Ingest Falabella",
    )
    db.add(dia)
    await db.flush()
    return dia


async def _ensure_vehicle(
    db: AsyncSession, empresa_id: int, patente_falsa: int, counter: list[int]
) -> int:
    plate = f"PAT-{patente_falsa:03d}"
    existing = (
        await db.execute(
            select(Vehicle).where(Vehicle.empresa_id == empresa_id, Vehicle.plate == plate)
        )
    ).scalar_one_or_none()
    if existing is not None:
        return existing.vehicle_id
    veh = Vehicle(
        empresa_id=empresa_id,
        nombre=plate,
        plate=plate,
        descripcion="Vehiculo sintetico generado por ingest Falabella",
        activo=True,
    )
    db.add(veh)
    await db.flush()
    counter[0] += 1
    return veh.vehicle_id


async def _ensure_driver_for_ruta(
    db: AsyncSession, empresa_id: int, idruta: int, counter: list[int]
) -> str:
    nombre = f"Driver Ruta {idruta}"
    existing = (
        await db.execute(
            select(Driver).where(Driver.empresa_id == empresa_id, Driver.nombre == nombre)
        )
    ).scalar_one_or_none()
    if existing is not None:
        return existing.driver_id

    # Next sequential driver_id.
    prefix = f"DRV-{empresa_id:02d}"
    n = int(
        (
            await db.execute(
                select(func.count(Driver.driver_id)).where(
                    Driver.driver_id.like(f"{prefix}%")
                )
            )
        ).scalar_one()
        or 0
    )
    new_id = f"{prefix}{n + 1:03d}"
    driver = Driver(
        driver_id=new_id,
        empresa_id=empresa_id,
        nombre=nombre,
        notify_whatsapp=False,
        activation_token=secrets.token_urlsafe(16),
        activo=True,
    )
    db.add(driver)
    await db.flush()
    counter[0] += 1
    return driver.driver_id


# ── Geocoding background task ───────────────────────────────────────────────


async def _geocode_clientes_for_empresas(empresa_ids: list[int]) -> None:
    """Eager geocode kick-off for empresas just ingested.

    CR-020: the heavy lifting is owned by the lifespan loop
    (`geocode_pending_clientes_loop`), but we still fire one immediate batch
    scoped to the empresas we just touched so the demo doesn't wait an interval
    cycle. A few hundred Nominatim calls at 1 req/s is fine — the loop will
    pick up whatever this leaves behind.
    """
    if not empresa_ids:
        return
    try:
        await geocode_pending_clientes(empresa_ids=empresa_ids, max_batch=500)
    except Exception as e:
        logger.warning(f"[geocode] post-ingest batch failed; loop will retry: {e}")


# ── Core pipeline ───────────────────────────────────────────────────────────


async def _ingest_rows(
    db: AsyncSession,
    raw_rows: list[dict[str, Any]],
    user: User,
) -> tuple[IngestResult, list[int]]:
    """Run the bulk pipeline. Returns (result, empresa_ids_for_bg_geocoding)."""
    t_pipeline = time.perf_counter()

    if not raw_rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No rows to process")
    logger.info(f"[ingest] received {len(raw_rows)} raw rows; validating…")

    # 1. Validate rows via Pydantic.
    t0 = time.perf_counter()
    rows: list[FalabellaRow] = []
    for i, raw in enumerate(raw_rows):
        try:
            rows.append(FalabellaRow.model_validate(raw))
        except Exception as e:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Row {i + 2} validation failed: {e}",
            ) from None
    logger.info(f"[ingest] validated {len(rows)} rows in {time.perf_counter() - t0:.2f}s")

    # 2. Sanity-check empresa mapping.
    unknown = sorted({r.empresa_falsa for r in rows if r.empresa_falsa not in EMPRESA_MAP})
    if unknown:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"empresa_falsa values out of mapping: {unknown}. Mapping: {EMPRESA_MAP}",
        )

    by_empresa: dict[int, list[FalabellaRow]] = defaultdict(list)
    for r in rows:
        by_empresa[EMPRESA_MAP[r.empresa_falsa]].append(r)
    logger.info(
        f"[ingest] {sum(len(v) for v in by_empresa.values())} rows across "
        f"{len(by_empresa)} empresa(s): "
        f"{ {k: len(v) for k, v in by_empresa.items()} }"
    )

    # 3. Verify target empresas exist.
    for emp_id in by_empresa:
        if (
            await db.execute(select(Empresa).where(Empresa.empresa_id == emp_id))
        ).scalar_one_or_none() is None:
            raise HTTPException(
                status.HTTP_500_INTERNAL_SERVER_ERROR,
                f"Target empresa_id={emp_id} not found. Run migration 0019.",
            )

    advertencias: list[str] = []
    dia_ids: list[int] = []
    rutas_creadas = 0
    visitas_creadas = 0
    clientes_creados = 0
    clientes_reusados = 0
    vehiculos_creados = [0]
    drivers_creados = [0]

    # CR-023 / CR-027: clientes are global (unique by RUT). Pre-fetch once
    # across all empresas in this ingest pass; per-empresa loops reuse this map.
    # CR-027 dropped the cliente_empresas accumulator — the link to a
    # transportista is *implicit* in the visitas rows we insert below.
    t0 = time.perf_counter()
    existing_clientes_global: dict[str, int] = {
        c.rut: c.cliente_id
        for c in (
            await db.execute(
                select(Cliente).where(Cliente.rut.like("FAL-%"))
            )
        ).scalars().all()
    }
    logger.info(
        f"[ingest] pre-fetched {len(existing_clientes_global)} global "
        f"FAL-* clientes in {time.perf_counter() - t0:.2f}s"
    )

    for empresa_id, emp_rows in by_empresa.items():
        t_emp = time.perf_counter()
        logger.info(f"[ingest] empresa {empresa_id}: {len(emp_rows)} rows")

        # ── 3a. Ensure dia ────────────────────────────────────────────
        fechas = [_parse_fechainicioruta(r.fechainicioruta).date() for r in emp_rows]
        fecha = min(fechas)
        dia = await _ensure_dia(db, empresa_id, fecha, user)
        dia_ids.append(dia.dia_id)
        logger.info(f"[ingest]   dia_id={dia.dia_id} estado={dia.estado} fecha={fecha}")

        if dia.estado not in ("BORRADOR", "VALIDADO"):
            advertencias.append(
                f"Dia {dia.dia_id} (empresa {empresa_id}, fecha {fecha}) en estado "
                f"{dia.estado}: solo se reusa cabecera, no se modifican visitas."
            )
            continue

        # ── 3b. Group rows by idruta ──────────────────────────────────
        by_ruta: dict[int, list[FalabellaRow]] = defaultdict(list)
        for r in emp_rows:
            by_ruta[r.idruta].append(r)

        # ── 3c. Ensure vehicles + drivers (low-volume: ~22 each) ──────
        t0 = time.perf_counter()
        ruta_meta: dict[int, dict[str, Any]] = {}
        for orden, (idruta, ruta_rows) in enumerate(by_ruta.items(), start=1):
            patente = ruta_rows[0].patente_falsa
            vehicle_id = await _ensure_vehicle(db, empresa_id, patente, vehiculos_creados)
            driver_id = await _ensure_driver_for_ruta(
                db, empresa_id, idruta, drivers_creados
            )
            ruta_meta[idruta] = {
                "orden": orden,
                "vehicle_id": vehicle_id,
                "driver_id": driver_id,
            }
        await db.commit()
        logger.info(
            f"[ingest]   ensured {len(ruta_meta)} vehicles+drivers "
            f"(v_new={vehiculos_creados[0]}, d_new={drivers_creados[0]}) "
            f"in {time.perf_counter() - t0:.2f}s"
        )

        # ── 3d. Pre-fetch existing rutas for this dia ─────────────────
        existing_rutas: dict[str, int] = {
            r.folio: r.ruta_id
            for r in (
                await db.execute(
                    select(Ruta).where(Ruta.dia_id == dia.dia_id, Ruta.folio.is_not(None))
                )
            ).scalars().all()
        }
        logger.info(f"[ingest]   {len(existing_rutas)} rutas pre-existing for dia")

        # ── 3e. Bulk-insert NEW rutas ────────────────────────────────
        rutas_to_insert: list[dict[str, Any]] = []
        for idruta, meta in ruta_meta.items():
            folio = str(idruta)
            if folio in existing_rutas:
                continue
            rutas_to_insert.append(
                {
                    "dia_id": dia.dia_id,
                    "driver_id": meta["driver_id"],
                    "vehicle_id": meta["vehicle_id"],
                    "orden": meta["orden"],
                    "folio": folio,
                    "subfolio": None,
                    "notas": "Ingest Falabella",
                }
            )
        if rutas_to_insert:
            t0 = time.perf_counter()
            await _chunked_insert(db, Ruta, rutas_to_insert, "rutas")
            # Recover PKs by re-selecting.
            inserted_rutas = (
                await db.execute(
                    select(Ruta.folio, Ruta.ruta_id).where(
                        Ruta.dia_id == dia.dia_id, Ruta.folio.is_not(None)
                    )
                )
            ).all()
            existing_rutas = {folio: rid for folio, rid in inserted_rutas}
            rutas_creadas += len(rutas_to_insert)
            logger.info(
                f"[ingest]   created {len(rutas_to_insert)} rutas in "
                f"{time.perf_counter() - t0:.2f}s"
            )

        # idruta -> ruta_id map
        idruta_to_ruta_id: dict[int, int] = {
            int(folio): rid for folio, rid in existing_rutas.items()
        }

        # ── 3f. Use GLOBAL cliente map (CR-023) ───────────────────────
        # clientes are unique by RUT globally; the pre-fetched map above
        # covers all FAL-* ruts already in the DB. We only insert truly-new
        # ones here.
        existing_clientes = existing_clientes_global  # alias for readability

        # ── 3g. Bulk-insert NEW clientes (identity-only, CR-027) ─────
        # Collect unique (do, head_row) pairs.
        unique_dos: dict[str, FalabellaRow] = {}
        for r in emp_rows:
            do_key = str(r.do)
            if do_key not in unique_dos:
                unique_dos[do_key] = r

        clientes_to_insert: list[dict[str, Any]] = []
        for do, head in unique_dos.items():
            surrogate_rut = f"FAL-{do}"
            if surrogate_rut in existing_clientes:
                clientes_reusados += 1
                continue
            # CR-020: synchronously populate lat_default/lon_default with the
            # comuna centroid so the simulator can move drivers immediately.
            # The Nominatim upgrade happens later via the lifespan loop.
            cen = centroide_comuna(head.localidad) if head.localidad else None
            lat_def: float | None = None
            lon_def: float | None = None
            geo_status = "pending"
            if cen is not None:
                lat_def, lon_def = cen
                geo_status = "centroide_fallback"
            clientes_to_insert.append(
                {
                    # CR-027: clientes are identity-only — no empresa link on
                    # the master. The tenant relationship lives in visitas.
                    "nombre": f"Cliente {do[-6:]}",
                    "rut": surrogate_rut,
                    "direccion_default": head.direccion,
                    "comuna_default": head.localidad,
                    "region_default": head.region,
                    "lat_default": lat_def,
                    "lon_default": lon_def,
                    "geocoding_status": geo_status,
                    "geocoding_attempts": 0,
                    "es_vip": False,
                }
            )

        if clientes_to_insert:
            t0 = time.perf_counter()
            await _chunked_insert(db, Cliente, clientes_to_insert, "clientes")
            # Recover PKs by re-selecting only the surrogate ruts we just added.
            new_ruts = [c["rut"] for c in clientes_to_insert]
            # SQL Server has a 2100-param limit; chunk the IN clause.
            for j in range(0, len(new_ruts), 1000):
                chunk_ruts = new_ruts[j : j + 1000]
                rows_back = (
                    await db.execute(
                        select(Cliente.rut, Cliente.cliente_id).where(
                            Cliente.rut.in_(chunk_ruts),
                        )
                    )
                ).all()
                for rut_val, cid in rows_back:
                    existing_clientes[rut_val] = cid
            clientes_creados += len(clientes_to_insert)
            logger.info(
                f"[ingest]   created {len(clientes_to_insert)} clientes in "
                f"{time.perf_counter() - t0:.2f}s "
                f"(reused {clientes_reusados})"
            )

        # ── 3h. Pre-fetch existing visitas for this dia (by ruta_id, folio_cliente) ──
        t0 = time.perf_counter()
        existing_visitas: set[tuple[int, str]] = {
            (r_id, fc)
            for r_id, fc in (
                await db.execute(
                    select(Visita.ruta_id, Visita.folio_cliente).where(
                        Visita.dia_id == dia.dia_id,
                        Visita.folio_cliente.is_not(None),
                    )
                )
            ).all()
        }
        logger.info(
            f"[ingest]   {len(existing_visitas)} visitas pre-existing "
            f"(query {time.perf_counter() - t0:.2f}s)"
        )

        # ── 3i. Build NEW visitas payload ────────────────────────────
        visitas_to_insert: list[dict[str, Any]] = []
        for idruta, ruta_rows in by_ruta.items():
            ruta_id = idruta_to_ruta_id.get(idruta)
            if ruta_id is None:
                advertencias.append(f"Ruta idruta={idruta} no fue insertada; visitas omitidas")
                continue

            by_do: dict[str, list[FalabellaRow]] = defaultdict(list)
            for r in ruta_rows:
                by_do[str(r.do)].append(r)

            visita_orden = 0
            for do, do_rows in by_do.items():
                visita_orden += 1
                if (ruta_id, do) in existing_visitas:
                    continue
                head = do_rows[0]
                surrogate_rut = f"FAL-{do}"
                cliente_id = existing_clientes.get(surrogate_rut)
                cliente_nombre = f"Cliente {do[-6:]}"

                # CR-027: NO cliente_empresas accumulator. The link cliente ->
                # empresa is implicit in this visita's `empresa_id` (and via
                # dia_id -> dias_operativos.empresa_id).

                # CR-020: populate lat/lon synchronously with comuna centroid
                # so the simulator has coords from the moment the visita is
                # inserted. Nominatim fine-grained upgrade happens later via
                # `geocode_pending_clientes` (background loop) and is cascaded
                # to these visitas when their cliente upgrades.
                v_cen = centroide_comuna(head.localidad) if head.localidad else None
                v_lat: float | None = None
                v_lon: float | None = None
                if v_cen is not None:
                    v_lat, v_lon = v_cen

                visitas_to_insert.append(
                    {
                        "ruta_id": ruta_id,
                        "dia_id": dia.dia_id,
                        "empresa_id": empresa_id,
                        "orden": visita_orden,
                        "cliente_id": cliente_id,
                        "cliente_nombre": cliente_nombre,
                        "cliente_rut": None,
                        "cliente_telefono": None,
                        "direccion": (head.direccion or "Sin direccion")[:500],
                        "comuna": head.localidad or None,
                        "region": head.region or None,
                        "lat": v_lat,
                        "lon": v_lon,
                        "estado": _map_estado(head.estado),
                        "estado_fuente": str(head.estado) if head.estado else None,
                        "motivo": str(head.motivonoentrega) if head.motivonoentrega else None,
                        "motivo_comentario": (
                            str(head.comentarionoentrega) if head.comentarionoentrega else None
                        ),
                        "motivo_ia_sugerido": None,
                        "n_bultos": len(do_rows),
                        "referencia": None,
                        "es_vip": 0,
                        "notas": None,
                        "folio_cliente": do,
                        "subfolio_bulto": (
                            str(head.Suborden)
                            if len(do_rows) == 1 and head.Suborden is not None
                            else None
                        ),
                        "parent_order": (str(head.parentorder) if head.parentorder else None),
                        "tipo_documento": (str(head.tipodocumento) if head.tipodocumento else None),
                        "fecha_pactada": _parse_fechapactada(head.fechapactada),
                    }
                )

        # ── 3j. Bulk-insert visitas ──────────────────────────────────
        if visitas_to_insert:
            t0 = time.perf_counter()
            await _chunked_insert(db, Visita, visitas_to_insert, "visitas")
            visitas_creadas += len(visitas_to_insert)
            logger.info(
                f"[ingest]   created {len(visitas_to_insert)} visitas in "
                f"{time.perf_counter() - t0:.2f}s"
            )

        logger.info(
            f"[ingest] empresa {empresa_id} done in "
            f"{time.perf_counter() - t_emp:.2f}s"
        )

    # Final safety commit (vehicles/drivers/dia adjustments).
    await db.commit()

    elapsed = time.perf_counter() - t_pipeline
    logger.info(
        f"[ingest] DONE in {elapsed:.2f}s: "
        f"empresas={len(by_empresa)} rutas={rutas_creadas} visitas={visitas_creadas} "
        f"clientes_new={clientes_creados} clientes_reused={clientes_reusados} "
        f"vehicles_new={vehiculos_creados[0]} drivers_new={drivers_creados[0]}"
    )

    result = IngestResult(
        dia_ids=sorted(set(dia_ids)),
        empresas_procesadas=len(by_empresa),
        rutas_creadas=rutas_creadas,
        visitas_creadas=visitas_creadas,
        clientes_creados=clientes_creados,
        clientes_reusados=clientes_reusados,
        vehiculos_creados=vehiculos_creados[0],
        drivers_creados=drivers_creados[0],
        geocoding_en_progreso=True,
        advertencias=advertencias,
    )
    return result, sorted(by_empresa.keys())


# ── Endpoints ───────────────────────────────────────────────────────────────


@router.post(
    "/falabella-xlsx",
    operation_id="ingestFalabellaXlsx",
    response_model=IngestResult,
    status_code=status.HTTP_201_CREATED,
    summary="Ingest Falabella source feed from XLSX (sheet 'Geo' preferred).",
    responses={
        400: {"description": "Bad rows or unmapped empresa_falsa"},
        403: {"description": "Only falabella_admin/ops"},
    },
)
async def ingest_falabella_xlsx(
    file: UploadFile,
    user: User = Depends(current_user),
    db: AsyncSession = Depends(get_db),
) -> IngestResult:
    _require_falabella_role(user)
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Only .xlsx supported")
    content = await file.read()
    logger.info(f"[ingest] received XLSX {file.filename} ({len(content)} bytes)")
    t0 = time.perf_counter()
    raw_rows = _parse_xlsx(content)
    logger.info(
        f"[ingest] parsed {len(raw_rows)} rows from XLSX in "
        f"{time.perf_counter() - t0:.2f}s"
    )
    result, empresa_ids = await _ingest_rows(db, raw_rows, user)
    # Fire-and-forget AFTER final commit; do NOT await.
    asyncio.create_task(_geocode_clientes_for_empresas(empresa_ids))
    return result


@router.post(
    "/falabella-json",
    operation_id="ingestFalabellaJson",
    response_model=IngestResult,
    status_code=status.HTTP_201_CREATED,
    summary="Ingest Falabella source feed from JSON body.",
)
async def ingest_falabella_json(
    body: FalabellaJSONBody,
    user: User = Depends(current_user),
    db: AsyncSession = Depends(get_db),
) -> IngestResult:
    _require_falabella_role(user)
    raw_rows = [r.model_dump() for r in body.rows]
    logger.info(f"[ingest] received {len(raw_rows)} JSON rows")
    result, empresa_ids = await _ingest_rows(db, raw_rows, user)
    asyncio.create_task(_geocode_clientes_for_empresas(empresa_ids))
    return result
