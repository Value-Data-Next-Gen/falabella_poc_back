"""Bulk import from XLSX — per-entity with downloadable templates.

Endpoints:
  GET  /api/v1/templates/{entity}                      → download XLSX template
  POST /api/v1/empresas/cargar-excel                   → import empresas (admin)
  POST /api/v1/empresas/{id}/conductores/cargar-excel   → import conductores for empresa
  POST /api/v1/empresas/{id}/vehiculos/cargar-excel     → import vehiculos for empresa
  POST /api/v1/empresas/{id}/contactos/cargar-excel     → import contactos for empresa
"""
from __future__ import annotations

import io
import secrets
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import current_user, require_admin
from app.core.security.scope import can_access_empresa
from app.db.models.driver import Driver
from app.db.models.empresa import Empresa
from app.db.models.empresa_contacto import EmpresaContacto
from app.db.models.user import User
from app.db.models.vehicle import Vehicle
from app.db.session import get_db

router = APIRouter(tags=["bulk-import"])

# ── Response models ──

class RowResult(BaseModel):
    fila: int
    estado: str
    detalle: str | None = None

class ImportResult(BaseModel):
    creados: int
    fallidos: int
    filas: list[RowResult]

# ── Templates ──

class _Col:
    __slots__ = ("header", "desc", "required", "example")
    def __init__(self, header: str, desc: str, required: bool, example: str):
        self.header = header
        self.desc = desc
        self.required = required
        self.example = example


TEMPLATES: dict[str, list[_Col]] = {
    "empresas": [
        _Col("nombre", "Nombre comercial de la empresa de transporte", True, "Falabella Transporte SpA"),
        _Col("rut", "RUT de la empresa con digito verificador", True, "76.823.145-1"),
        _Col("razon_social", "Razon social completa", False, "Transportes Falabella Servicios Logisticos SpA"),
        _Col("region", "Region de Chile donde opera (ej: Metropolitana de Santiago)", False, "Metropolitana de Santiago"),
        _Col("comuna", "Comuna principal de operacion", False, "Santiago"),
        _Col("central_phone", "Telefono central en formato E.164", False, "+56226234500"),
    ],
    "conductores": [
        _Col("nombre", "Nombre completo del conductor", True, "Ricardo Diaz Espinoza"),
        _Col("telefono", "Telefono celular en formato E.164 (con +56)", False, "+56954861116"),
        _Col("rut", "RUT del conductor con digito verificador", False, "12.345.678-9"),
        _Col("patente_vehiculo", "Patente del vehiculo a asignar (debe existir en la empresa)", False, "LKKG-23"),
    ],
    "vehiculos": [
        _Col("nombre", "Codigo interno del vehiculo", True, "FUR-0101"),
        _Col("patente", "Patente unica del vehiculo", True, "LKKG-23"),
        _Col("tipo", "Tipo: Furgon Pequeno, Furgon Mediano, Furgon Grande, Camion Liviano, Camion Mediano, Camion 3/4", False, "Furgon Mediano"),
        _Col("capacidad_m3", "Capacidad de carga en metros cubicos (numero entero)", False, "15"),
        _Col("descripcion", "Descripcion adicional del vehiculo (marca, modelo, color, etc.)", False, "Hyundai Porter 2024 blanco"),
    ],
    "contactos": [
        _Col("nombre", "Nombre completo del contacto", True, "Carolina Castro Munoz"),
        _Col("rol", "Rol: jefe, coordinador, otro", True, "jefe"),
        _Col("telefono", "Telefono celular en formato E.164 (con +56)", False, "+56974232322"),
        _Col("email", "Correo electronico", False, "carolina@empresa.cl"),
    ],
}


@router.get(
    "/api/v1/templates/{entity}",
    operation_id="downloadTemplate",
    summary="Descarga plantilla XLSX para carga masiva.",
)
async def download_template(entity: str) -> StreamingResponse:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    cols = TEMPLATES.get(entity)
    if not cols:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"Plantilla no encontrada: {entity}")

    wb = Workbook()

    # ── Instructions sheet ──
    ws_info = wb.active
    assert ws_info is not None
    ws_info.title = "Instrucciones"
    ws_info.sheet_properties.tabColor = "1F4E79"

    info_rows = [
        ("PLANTILLA DE CARGA MASIVA", ""),
        ("", ""),
        ("Entidad:", entity.capitalize()),
        ("", ""),
        ("INSTRUCCIONES:", ""),
        ("1.", "Complete la hoja 'Datos' con la informacion solicitada."),
        ("2.", "Los campos marcados con (*) son obligatorios."),
        ("3.", "No modifique la fila de encabezados (fila 1 de la hoja Datos)."),
        ("4.", "Puede eliminar la fila de ejemplo (fila 3) antes de cargar."),
        ("5.", "Formatos de telefono: siempre con codigo pais, ej: +56912345678"),
        ("6.", "Guarde el archivo como .xlsx y carguelo en la plataforma."),
        ("", ""),
        ("CAMPOS:", ""),
    ]
    for col_def in cols:
        req = "(*) OBLIGATORIO" if col_def.required else "Opcional"
        info_rows.append((f"  {col_def.header}", f"{col_def.desc}. {req}. Ejemplo: {col_def.example}"))

    title_font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    header_font = Font(name="Arial", size=11, bold=True)
    body_font = Font(name="Arial", size=11)

    for r, (a, b) in enumerate(info_rows, start=1):
        ca = ws_info.cell(row=r, column=1, value=a)
        cb = ws_info.cell(row=r, column=2, value=b)
        ca.font = title_font if r == 1 else header_font if a.endswith(":") else body_font
        cb.font = body_font
        cb.alignment = Alignment(wrap_text=True)

    ws_info.column_dimensions["A"].width = 20
    ws_info.column_dimensions["B"].width = 80

    # ── Data sheet ──
    ws = wb.create_sheet(title="Datos")
    wb.active = wb.index(ws)

    green_fill = PatternFill("solid", fgColor="159349")
    light_fill = PatternFill("solid", fgColor="E8F5ED")
    desc_fill = PatternFill("solid", fgColor="F5F5F5")
    white_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    desc_font = Font(name="Arial", size=9, italic=True, color="5B6776")
    example_font = Font(name="Arial", size=11, color="8B98A9")
    thin_border = Border(
        left=Side(style="thin", color="DADADA"),
        right=Side(style="thin", color="DADADA"),
        top=Side(style="thin", color="DADADA"),
        bottom=Side(style="thin", color="DADADA"),
    )

    col_letters = [chr(65 + i) for i in range(len(cols))]

    for i, col_def in enumerate(cols):
        letter = col_letters[i]
        ci = i + 1

        # Row 1: header
        label = f"{col_def.header} *" if col_def.required else col_def.header
        cell_h = ws.cell(row=1, column=ci, value=label.upper())
        cell_h.font = white_font
        cell_h.fill = green_fill
        cell_h.border = thin_border
        cell_h.alignment = Alignment(horizontal="center")

        # Row 2: description
        cell_d = ws.cell(row=2, column=ci, value=col_def.desc)
        cell_d.font = desc_font
        cell_d.fill = desc_fill
        cell_d.border = thin_border
        cell_d.alignment = Alignment(wrap_text=True)

        # Row 3: example
        cell_e = ws.cell(row=3, column=ci, value=col_def.example)
        cell_e.font = example_font
        cell_e.fill = light_fill
        cell_e.border = thin_border

        ws.column_dimensions[letter].width = max(25, len(col_def.desc) // 3)

    ws.row_dimensions[2].height = 35
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="plantilla_{entity}.xlsx"'},
    )

# ── Helpers ──

def _parse_sheet(ws: Any) -> list[dict[str, Any]]:
    """Parse XLSX sheet. Row 1 = headers, row 2 = descriptions (skipped), row 3+ = data."""
    headers: list[str] = []
    rows: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [
                str(h).strip().lower().replace(" *", "").replace("*", "") if h else f"col_{j}"
                for j, h in enumerate(row)
            ]
            continue
        if i == 1:
            continue
        if all(c is None for c in row):
            continue
        rows.append(dict(zip(headers, row, strict=False)))
    return rows

def _cell(row: dict[str, Any], *keys: str) -> str | None:
    for k in keys:
        val = row.get(k) or row.get(k.lower())
        if val is not None:
            return str(val).strip()
    return None

async def _parse_upload(file: UploadFile) -> list[dict[str, Any]]:
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Solo archivos .xlsx")
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"No se pudo leer el archivo: {e}") from None
    ws = wb.active
    rows = _parse_sheet(ws)
    wb.close()
    return rows

# ── Empresas bulk import (admin only) ──

@router.post(
    "/api/v1/empresas/cargar-excel",
    operation_id="bulkImportEmpresas",
    response_model=ImportResult,
    dependencies=[Depends(require_admin())],
    summary="Carga masiva de empresas desde XLSX.",
)
async def bulk_import_empresas(
    file: UploadFile,
    db: AsyncSession = Depends(get_db),
) -> ImportResult:
    rows = await _parse_upload(file)
    results: list[RowResult] = []
    created = 0
    for i, row in enumerate(rows, start=2):
        nombre = _cell(row, "nombre")
        rut = _cell(row, "rut")
        if not nombre or not rut:
            results.append(RowResult(fila=i, estado="error", detalle="nombre y rut son obligatorios"))
            continue
        empresa = Empresa(
            nombre=nombre, rut=rut,
            razon_social=_cell(row, "razon_social"),
            region=_cell(row, "region"),
            comuna=_cell(row, "comuna"),
            central_phone=_cell(row, "central_phone", "telefono"),
        )
        try:
            # SAVEPOINT per row: a duplicate/invalid row rolls back ONLY itself.
            # A bare db.rollback() here would discard every previously-flushed
            # row in the batch while still reporting them as "creado".
            async with db.begin_nested():
                db.add(empresa)
                await db.flush()
            results.append(RowResult(fila=i, estado="creado"))
            created += 1
        except IntegrityError as e:
            results.append(RowResult(fila=i, estado="error", detalle=str(e)[:200]))
    await db.commit()
    return ImportResult(creados=created, fallidos=len(rows) - created, filas=results)

# ── Conductores bulk import (scoped to empresa) ──

@router.post(
    "/api/v1/empresas/{empresa_id}/conductores/cargar-excel",
    operation_id="bulkImportConductores",
    response_model=ImportResult,
    summary="Carga masiva de conductores para una empresa.",
)
async def bulk_import_conductores(
    empresa_id: int,
    file: UploadFile,
    user: User = Depends(current_user),
    db: AsyncSession = Depends(get_db),
) -> ImportResult:
    if not can_access_empresa(user, empresa_id):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Fuera de alcance")
    rows = await _parse_upload(file)
    results: list[RowResult] = []
    created = 0
    # Next driver-id sequence computed ONCE before the loop. The old code ran a
    # full-table scan per row (`len(.all())`) and minted DRV-{empresa}{count+1},
    # which (a) was O(n²) and (b) collided on delete-gaps (count < max id).
    # Taking MAX of the existing numeric suffixes is correct under deletes.
    prefix = f"DRV-{empresa_id:02d}"
    existing_ids = (
        await db.execute(select(Driver.driver_id).where(Driver.empresa_id == empresa_id))
    ).scalars().all()
    seq = max(
        (int(d[len(prefix):]) for d in existing_ids
         if d and d.startswith(prefix) and d[len(prefix):].isdigit()),
        default=0,
    )
    for i, row in enumerate(rows, start=2):
        nombre = _cell(row, "nombre")
        if not nombre:
            results.append(RowResult(fila=i, estado="error", detalle="nombre es obligatorio"))
            continue
        vehicle_id = None
        patente = _cell(row, "patente_vehiculo", "patente", "plate")
        if patente:
            veh_result = await db.execute(
                select(Vehicle).where(Vehicle.plate == patente, Vehicle.empresa_id == empresa_id)
            )
            veh = veh_result.scalar_one_or_none()
            if veh:
                vehicle_id = veh.vehicle_id
            else:
                results.append(RowResult(fila=i, estado="error", detalle=f"Vehiculo con patente '{patente}' no encontrado en esta empresa"))
                continue

        seq += 1
        driver_id = f"{prefix}{seq:03d}"
        driver = Driver(
            driver_id=driver_id, empresa_id=empresa_id, nombre=nombre,
            phone_e164=_cell(row, "telefono", "phone_e164"),
            rut=_cell(row, "rut"),
            vehicle_id=vehicle_id,
            activation_token=secrets.token_urlsafe(16),
            notify_whatsapp=True,
        )
        try:
            # SAVEPOINT per row (see empresas importer for rationale).
            async with db.begin_nested():
                db.add(driver)
                await db.flush()
            results.append(RowResult(fila=i, estado="creado", detalle=driver_id))
            created += 1
        except IntegrityError as e:
            results.append(RowResult(fila=i, estado="error", detalle=str(e)[:200]))
    await db.commit()
    return ImportResult(creados=created, fallidos=len(rows) - created, filas=results)

# ── Vehiculos bulk import (scoped to empresa) ──

@router.post(
    "/api/v1/empresas/{empresa_id}/vehiculos/cargar-excel",
    operation_id="bulkImportVehiculos",
    response_model=ImportResult,
    summary="Carga masiva de vehiculos para una empresa.",
)
async def bulk_import_vehiculos(
    empresa_id: int,
    file: UploadFile,
    user: User = Depends(current_user),
    db: AsyncSession = Depends(get_db),
) -> ImportResult:
    if not can_access_empresa(user, empresa_id):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Fuera de alcance")
    rows = await _parse_upload(file)
    results: list[RowResult] = []
    created = 0
    for i, row in enumerate(rows, start=2):
        nombre = _cell(row, "nombre")
        plate = _cell(row, "patente", "plate")
        if not nombre or not plate:
            results.append(RowResult(fila=i, estado="error", detalle="nombre y patente son obligatorios"))
            continue
        cap = _cell(row, "capacidad_m3", "capacity_m3")
        vehicle = Vehicle(
            empresa_id=empresa_id, nombre=nombre, plate=plate,
            tipo=_cell(row, "tipo"),
            capacity_m3=int(cap) if cap and cap.isdigit() else None,
            descripcion=_cell(row, "descripcion"),
        )
        try:
            # SAVEPOINT per row (see empresas importer for rationale).
            async with db.begin_nested():
                db.add(vehicle)
                await db.flush()
            results.append(RowResult(fila=i, estado="creado"))
            created += 1
        except IntegrityError as e:
            results.append(RowResult(fila=i, estado="error", detalle=str(e)[:200]))
    await db.commit()
    return ImportResult(creados=created, fallidos=len(rows) - created, filas=results)

# ── Contactos bulk import (scoped to empresa) ──

@router.post(
    "/api/v1/empresas/{empresa_id}/contactos/cargar-excel",
    operation_id="bulkImportContactos",
    response_model=ImportResult,
    summary="Carga masiva de contactos para una empresa.",
)
async def bulk_import_contactos(
    empresa_id: int,
    file: UploadFile,
    user: User = Depends(current_user),
    db: AsyncSession = Depends(get_db),
) -> ImportResult:
    if not can_access_empresa(user, empresa_id):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Fuera de alcance")
    rows = await _parse_upload(file)
    results: list[RowResult] = []
    created = 0
    valid_roles = {"jefe", "coordinador", "otro"}
    for i, row in enumerate(rows, start=2):
        nombre = _cell(row, "nombre")
        rol = _cell(row, "rol")
        if not nombre or not rol:
            results.append(RowResult(fila=i, estado="error", detalle="nombre y rol son obligatorios"))
            continue
        if rol not in valid_roles:
            results.append(RowResult(fila=i, estado="error", detalle=f"rol invalido: {rol}. Usar: {', '.join(sorted(valid_roles))}"))
            continue
        contacto = EmpresaContacto(
            empresa_id=empresa_id, nombre=nombre, rol=rol,
            phone_e164=_cell(row, "telefono", "phone_e164"),
            email=_cell(row, "email"),
            activation_token=secrets.token_urlsafe(16),
        )
        try:
            # SAVEPOINT per row (see empresas importer for rationale).
            async with db.begin_nested():
                db.add(contacto)
                await db.flush()
            results.append(RowResult(fila=i, estado="creado"))
            created += 1
        except IntegrityError as e:
            results.append(RowResult(fila=i, estado="error", detalle=str(e)[:200]))
    await db.commit()
    return ImportResult(creados=created, fallidos=len(rows) - created, filas=results)
